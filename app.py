import os
from dotenv import load_dotenv
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, jsonify, make_response, current_app
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.urls import url_parse
from datetime import datetime, timedelta, timezone
import pandas as pd
import os
from werkzeug.utils import secure_filename
import tempfile
import re
from flask_migrate import Migrate
import requests
import msal
import secrets
import json
from flask_wtf import FlaskForm
import base64
from io import BytesIO
import os
from dotenv import load_dotenv
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import pandas as pd
import os
from werkzeug.utils import secure_filename
import tempfile
import re
from flask_migrate import Migrate
import requests
import msal
import secrets
import json
from flask_wtf import FlaskForm
import base64
from io import BytesIO
from openpyxl.utils import get_column_letter
from math import ceil
import uuid
import shutil
from flask_wtf.csrf import CSRFProtect, generate_csrf
from openpyxl import Workbook
import io
import pytz
from urllib.parse import quote_plus
import ldap3
from ldap3 import Server, Connection, ALL, NTLM, SUBTREE
from dateutil import parser as date_parser
# Mevcut importlarınıza bunları ekleyin:
from sqlalchemy import text
from sqlalchemy.exc import SQLAlchemyError
# Load environment variables
load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secret-key')

# SQL Server Configuration
from urllib.parse import quote_plus

# Azure SQL Server Configuration (for production)
# Use environment variables if available, otherwise use defaults
params = {
    'DRIVER': os.getenv('DB_DRIVER', '{ODBC Driver 17 for SQL Server}'),
    'SERVER': os.getenv('DB_SERVER', 'pmbomft.database.windows.net'),
    'DATABASE': os.getenv('DB_NAME', 'PMboMft'),
    'UID': os.getenv('DB_USER', 'mbomft_admin'),
    'PWD': os.getenv('DB_PASSWORD', 'Pp123456Pp123456'),
    'Encrypt': 'yes',
    'TrustServerCertificate': 'no',
    'Connection Timeout': '30'
}

# Local SQL Server Configuration (for development)
# Uncomment below and comment above for local development
# params = {
#     'DRIVER': '{ODBC Driver 17 for SQL Server}',
#     'SERVER': 'DESKTOP-E62JOKI\SQLEXPRESS',
#     'DATABASE': 'mbo-mft',
#     'UID': 'mft_user',
#     'PWD': '123456'
# }

conn_str = ';'.join(f"{k}={v}" for k, v in params.items())

# SQL Server connection string with SQL Authentication
app.config['SQLALCHEMY_DATABASE_URI'] = f"mssql+pyodbc:///?odbc_connect={quote_plus(conn_str)}"
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 3600,
    'pool_size': 10,
    'max_overflow': 20
}

# Add error handling for database connection
try:
    print(f"Attempting to connect to database: {params['SERVER']}")
    print(f"Database: {params['DATABASE']}")
    print(f"User: {params['UID']}")
except Exception as e:
    print(f"Database configuration error: {e}")

# Ensure uploads directory exists and has proper permissions
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Create uploads directory if it doesn't exist
if not os.path.exists(UPLOAD_FOLDER):
    try:
        os.makedirs(UPLOAD_FOLDER)
        print(f"Created uploads directory at: {UPLOAD_FOLDER}")
    except Exception as e:
        print(f"Error creating uploads directory: {str(e)}")
else:
    print(f"Uploads directory already exists at: {UPLOAD_FOLDER}")

app.config['TEMPLATES_AUTO_RELOAD'] = True
db = SQLAlchemy(app)
migrate = Migrate(app, db)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
csrf = CSRFProtect(app)

# Ensure upload folder exists
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# Microsoft OAuth Configuration
AZURE_CLIENT_ID = '6071b05c-29be-4447-9970-81cf9fbe88e4'
AZURE_CLIENT_SECRET = os.getenv('AZURE_CLIENT_SECRET')
AZURE_TENANT_ID = 'd0a51ea4-f656-42f7-a1b0-b70bb1c08731'
AZURE_AUTHORITY = f"https://login.microsoftonline.com/{AZURE_TENANT_ID}"
AZURE_SCOPES = ["User.Read", "User.Read.All", "User.ReadBasic.All"]

# Allowed Microsoft email addresses
ALLOWED_MS_EMAILS = [
    'ozer44@hotmail.com',  # Added new email
]

# Define both possible redirect URIs
LOCAL_IP = '192.168.1.26'
REDIRECT_URIS = [
    f"https://{LOCAL_IP}/login/microsoft/callback",
    "https://localhost/login/microsoft/callback",
    "https://127.0.0.1/login/microsoft/callback"
]
REDIRECT_URI = REDIRECT_URIS[0]  # Use local IP by default

# Initialize MSAL application
msal_app = msal.ConfidentialClientApplication(
    AZURE_CLIENT_ID,
    authority=AZURE_AUTHORITY,
    client_credential=AZURE_CLIENT_SECRET
)

# JSON İçe Aktarma için hedef alanlar
TARGET_FIELDS = [
    'Scenario Name',
    'Recurrence',
    'Type',
    'Action',
    'Source',
    'Destination',
    'User',
    'Email Notification'
]

TURKEY_TZ = timezone(timedelta(hours=3))
def turkey_now():
    return datetime.now(TURKEY_TZ)

class Role(db.Model):
    __tablename__ = 'roles'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    description = db.Column(db.String(200))
    is_admin = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=turkey_now)
    updated_at = db.Column(
        db.DateTime,
        default=turkey_now,
        onupdate=turkey_now)
    permissions = db.relationship(
        'Permission',
        secondary='role_permissions',
        lazy='dynamic')


class Permission(db.Model):
    __tablename__ = 'permissions'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    description = db.Column(db.String(200))

    def __init__(self, name, description=None):
        self.name = name
        self.description = description

    def __repr__(self):
        return f'<Permission {self.name}>'


class RolePermission(db.Model):
    __tablename__ = 'role_permissions'
    role_id = db.Column(
        db.Integer,
        db.ForeignKey('roles.id'),
        primary_key=True)
    permission_id = db.Column(
        db.Integer,
        db.ForeignKey('permissions.id'),
        primary_key=True)
    created_at = db.Column(db.DateTime, default=turkey_now)


class UserRole(db.Model):
    __tablename__ = 'user_roles'
    user_id = db.Column(
        db.Integer,
        db.ForeignKey('users.id'),
        primary_key=True)
    role_id = db.Column(
        db.Integer,
        db.ForeignKey('roles.id'),
        primary_key=True)
    created_at = db.Column(db.DateTime, default=turkey_now)


class User(UserMixin, db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(64), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(128))
    ms_id = db.Column(db.String(100), nullable=True)  # Removed unique=True
    is_oauth_user = db.Column(db.Boolean, default=False)
    is_ad_user = db.Column(db.Boolean, default=False)  # Added is_ad_user field
    is_admin = db.Column(db.Boolean, default=False)
    full_name = db.Column(db.String(100))
    last_login = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=turkey_now)
    default_landing_page = db.Column(db.String(50), default='index')
    theme = db.Column(db.String(20), default='light')
    sidebar_collapsed = db.Column(db.Boolean, default=False)
    show_email_notifications = db.Column(db.Boolean, default=True)
    items_per_page = db.Column(db.Integer, default=10)
    profile_photo = db.Column(db.Text)
    role_id = db.Column(db.Integer, db.ForeignKey('roles.id'))
    role = db.relationship('Role', backref=db.backref('users', lazy='dynamic'))

    # Many-to-many relationship with roles
    roles = db.relationship('Role', secondary='user_roles',
                           backref=db.backref('role_users', lazy='dynamic'),
                           overlaps="role,users")

    def has_permission(self, permission_name):
        """Check if user has a specific permission through their roles."""
        if self.is_admin:
            return True
        for role in self.roles:
            for permission in role.permissions:
                if permission.name == permission_name:
                    return True
        return False

    def set_password(self, password):
        """Generate a password hash for the provided password."""
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        """Check if the provided password matches the stored hash."""
        if not self.password_hash:
            return False
        return check_password_hash(self.password_hash, password)


class SystemSettings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    system_name = db.Column(
        db.String(100),
        nullable=False,
        default='MFT Tool Scenarios')
    timezone = db.Column(
        db.String(50),
        nullable=False,
        default='Europe/Istanbul')
    date_format = db.Column(
        db.String(20),
        nullable=False,
        default='DD/MM/YYYY')
    min_password_length = db.Column(db.Integer, nullable=False, default=8)
    session_timeout = db.Column(db.Integer, nullable=False, default=30)
    two_factor = db.Column(db.Boolean, default=False)
    backup_frequency = db.Column(
        db.String(20),
        nullable=False,
        default='daily')
    backup_location = db.Column(
        db.String(200),
        nullable=False,
        default='/backups')
    backup_retention = db.Column(db.Integer, nullable=False, default=30)
    log_level = db.Column(db.String(20), nullable=False, default='INFO')
    log_path = db.Column(
        db.String(200),
        nullable=False,
        default='/logs/app.log')
    log_rotation = db.Column(db.Boolean, default=True)
    # False = both, True = microsoft_only
    login_method = db.Column(db.Boolean, default=False)
    # Active Directory Settings
    ad_enabled = db.Column(db.Boolean, default=False)
    ad_server = db.Column(db.String(200))
    ad_domain = db.Column(db.String(200))
    ad_base_dn = db.Column(db.String(200))
    ad_username = db.Column(db.String(200))
    ad_password = db.Column(db.String(200))
    ad_port = db.Column(db.Integer, default=389)
    ad_use_ssl = db.Column(db.Boolean, default=False)
    ad_search_base = db.Column(db.String(200))  # Yeni alan
    ad_search_filter = db.Column(db.String(200), default='(sAMAccountName={username})')
    ad_group_filter = db.Column(db.String(200), default='(memberOf={group})')
    ad_admin_group = db.Column(db.String(200), default='CN=Administrators,CN=Builtin,DC=domain,DC=com')


class PaginatedGroups:
    def __init__(self, items, page, per_page, total):
        self._items = items
        self.page = page
        self.per_page = per_page
        self.total = total

    @property
    def items(self):
        return self._items

    @property
    def pages(self):
        return (self.total + self.per_page - 1) // self.per_page

    @property
    def has_prev(self):
        return self.page > 1

    @property
    def has_next(self):
        return self.page < self.pages

    @property
    def prev_num(self):
        return self.page - 1

    @property
    def next_num(self):
        return self.page + 1

    def iter_pages(self, left_edge=2, left_current=2, right_current=5, right_edge=2):
        last = 0
        for num in range(1, self.pages + 1):
            if (num <= left_edge or
                (num > self.page - left_current - 1 and
                 num < self.page + right_current) or
                num > self.pages - right_edge + 1):
                if last + 1 != num:
                    yield None
                yield num
                last = num


class Scenario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    custom_id = db.Column(db.String(100), nullable=True)
    scenario_name = db.Column(db.String(500), nullable=False)  # Increased from 200 to 500
    type_of_action = db.Column(db.String(50), nullable=False)
    user = db.Column(db.String(100), nullable=False)
    recurrence = db.Column(db.String(50), nullable=False)
    repeat_rate = db.Column(db.String(100), nullable=True)
    source = db.Column(db.String(1000), nullable=False)  # Increased from 500 to 1000
    destination = db.Column(db.String(1000), nullable=False)  # Increased from 500 to 1000
    email_notification = db.Column(db.Text, nullable=True)  # Changed to Text type
    responsible_person = db.Column(db.String(100), nullable=False)
    created_at = db.Column(db.DateTime, default=turkey_now)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    status = db.Column(db.String(20), default='active')
    group_id = db.Column(
        db.String(36),
        nullable=False,
        default=lambda: str(
            uuid.uuid4()))
    folder_id = db.Column(db.String(100), nullable=True)
    type = db.Column(db.String(50), nullable=True)
    step_count = db.Column(db.Integer, nullable=True)
    department = db.Column(db.String(100), nullable=True)

    def __init__(self, **kwargs):
        if 'group_id' not in kwargs:
            kwargs['group_id'] = str(uuid.uuid4())
        super(Scenario, self).__init__(**kwargs)


class ChangeRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    scenario_id = db.Column(db.Integer, db.ForeignKey('scenario.id'))
    requested_by = db.Column(db.Integer, db.ForeignKey(
        'users.id'))  # Changed to match the users table
    description = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(20), default='pending')
    created_at = db.Column(db.DateTime, default=turkey_now)
    field_to_change = db.Column(db.String(100))
    current_value = db.Column(db.String(500))
    requested_value = db.Column(db.String(500))

    # Add relationship to Scenario model
    scenario = db.relationship(
        'Scenario', backref=db.backref(
            'change_requests', lazy=True))
    user = db.relationship(
        'User',
        backref=db.backref(
            'change_requests',
            lazy=True))  # Updated relationship


class AllowedEmail(db.Model):
    __tablename__ = 'allowed_email'
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), nullable=False, unique=True)
    # Changed from 'user.id' to 'users.id'
    added_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=turkey_now)
    is_active = db.Column(db.Boolean, default=True)
    is_admin = db.Column(db.Boolean, default=False)


class Log(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    scenario_id = db.Column(
        db.Integer,
        db.ForeignKey('scenario.id'),
        nullable=True)  # Optional scenario association
    timestamp = db.Column(db.DateTime, default=turkey_now)
    log_type = db.Column(db.String(50))  # 'manual', 'auto', 'system'
    level = db.Column(db.String(20))  # 'INFO', 'ERROR', 'WARNING'
    message = db.Column(db.Text)
    source = db.Column(db.String(200))  # File path or manual entry
    parsed_data = db.Column(db.JSON)  # Parsed and analyzed log data
    created_by = db.Column(db.Integer, db.ForeignKey(
        'users.id'))  # Changed to match the users table
    # 'general', 'security', 'performance', 'error', etc.
    category = db.Column(db.String(50), default='general')
    host = db.Column(db.String(100))  # Source host/system name
    process = db.Column(db.String(100))  # Process or component name

    # Relationships
    scenario = db.relationship(
        'Scenario', backref=db.backref(
            'logs', lazy=True))
    user = db.relationship(
        'User', backref=db.backref(
            'logs', lazy=True))  # Updated relationship


class LogSettings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    auto_import_enabled = db.Column(db.Boolean, default=False)
    import_directory = db.Column(db.String(500))
    file_pattern = db.Column(db.String(100))  # e.g., "*.log"
    parse_interval = db.Column(db.Integer, default=5)  # minutes
    retention_period = db.Column(db.Integer, default=30)  # days
    max_file_size = db.Column(db.Integer, default=10)  # MB
    created_at = db.Column(db.DateTime, default=turkey_now)
    updated_at = db.Column(db.DateTime, onupdate=turkey_now)

class AuthLog(db.Model):
    __tablename__ = 'auth_logs'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), nullable=False)
    action = db.Column(db.String(50), nullable=False)  # login, logout, failed_login, role_assigned, role_revoked, scenario_updated, responsible_updated
    status = db.Column(db.String(20), nullable=False)  # success, failed
    ip_address = db.Column(db.String(50), nullable=False)
    timestamp = db.Column(db.DateTime, default=turkey_now)  # Always store in UTC
    details = db.Column(db.String(1000), nullable=True)
    user_agent = db.Column(db.String(500), nullable=True)
    target_user = db.Column(db.String(100))  # İşlemin yapıldığı kullanıcı
    target_role = db.Column(db.String(100))  # İşlemin yapıldığı rol
    target_scenario = db.Column(db.String(500))  # İşlemin yapıldığı senaryo
    performed_by = db.Column(db.String(100))  # İşlemi yapan kullanıcı

    @property
    def turkey_time(self):
        """Convert UTC timestamp to Turkey time"""
        if self.timestamp:
            turkey_tz = pytz.timezone('Europe/Istanbul')
            return self.timestamp.replace(tzinfo=pytz.UTC).astimezone(turkey_tz)
        return None


class ResponsiblePerson(db.Model):
    __tablename__ = 'responsible_persons'
    
    id = db.Column(db.Integer, primary_key=True)
    scenario_name = db.Column(db.String(500), nullable=False)  # Senaryo adı
    responsible_name = db.Column(db.String(100), nullable=False)  # Sorumlu kişi
    is_active = db.Column(db.Boolean, default=True)  # Senaryo aktif mi?
    created_at = db.Column(db.DateTime, default=turkey_now)  # Kayıt tarihi
    updated_at = db.Column(db.DateTime, default=turkey_now, onupdate=turkey_now)  # Güncelleme tarihi

    def __repr__(self):
        return f'<ResponsiblePerson {self.scenario_name}>'


class RestApiConnection(db.Model):
    __tablename__ = 'rest_api_connections'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    ip_address = db.Column(db.String(100), nullable=False)
    port = db.Column(db.Integer, nullable=False)
    username = db.Column(db.String(100), nullable=False)
    password = db.Column(db.String(100), nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    auth_token = db.Column(db.String(500), nullable=True)
    last_connection = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=turkey_now)
    updated_at = db.Column(db.DateTime, default=turkey_now, onupdate=turkey_now)
    site_id = db.Column(db.String(100), nullable=False, default='1')  # Default site ID

    @property
    def base_url(self):
        return f'http://{self.ip_address}:{self.port}'

    def __repr__(self):
        return f'<RestApiConnection {self.name}>'


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


@app.route('/health')
def health_check():
    """Health check endpoint for Azure"""
    try:
        # Test database connection - TEXT() içinde kullanın
        db.session.execute(text('SELECT 1'))
        return jsonify({
            'status': 'healthy',
            'database': 'connected',
            'server': params['SERVER'],
            'database_name': params['DATABASE']
        }), 200
    except SQLAlchemyError as e:
        return jsonify({
            'status': 'unhealthy',
            'database': 'disconnected',
            'error': str(e),
            'server': params['SERVER'],
            'database_name': params['DATABASE']
        }), 500
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'database': 'disconnected',
            'error': str(e),
            'server': params['SERVER'],
            'database_name': params['DATABASE']
        }), 500

@app.route('/')
@login_required
def index():
    # Get system settings
    system_settings = SystemSettings.query.first()
    
    # Get all scenarios
    scenarios = Scenario.query.all()
    
    # Get all users for responsible persons
    users = User.query.filter_by(is_active=True).all()
    responsible_persons = {user.full_name or user.username: user.full_name or user.username for user in users}
    
    # Create scenario groups
    scenario_groups = {}
    for scenario in scenarios:
        if scenario.scenario_name not in scenario_groups:
            scenario_groups[scenario.scenario_name] = {
                'scenarios': [],
                'action_types': set()
            }
        scenario_groups[scenario.scenario_name]['scenarios'].append(scenario)
        scenario_groups[scenario.scenario_name]['action_types'].add(scenario.type_of_action)
    
    # Count action types
    action_counts = {}
    for scenario in scenarios:
        if scenario.type_of_action not in action_counts:
            action_counts[scenario.type_of_action] = 0
        action_counts[scenario.type_of_action] += 1
    
    # Get search and filter parameters
    search = request.args.get('search', '')
    filter_type = request.args.get('filter_type', '')

    # Filter scenarios based on search and filter
    if search:
        search = search.strip().lower()
        filtered_groups = {}
        for name, group in scenario_groups.items():
            if (filter_type == 'name' and search in name.lower()) or \
               (filter_type == 'steps' and any(search in str(s.step_count).lower() for s in group['scenarios'])) or \
               (not filter_type and (search in name.lower() or 
                                   any(search in str(s.step_count).lower() for s in group['scenarios']))):
                filtered_groups[name] = group
        scenario_groups = filtered_groups
    
    # Calculate related scenarios count
    related_scenarios_count = {}
    for scenario in scenarios:
        base_name = get_base_scenario_name(scenario.scenario_name)
        if base_name not in related_scenarios_count:
            related_scenarios_count[base_name] = 0
        related_scenarios_count[base_name] += 1
    
    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = 10
    total = len(scenario_groups)
    
    # Calculate start and end indices for current page
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    
    # Get items for current page
    current_page_items = dict(list(scenario_groups.items())[start_idx:end_idx])
    
    # Create pagination object
    pagination = PaginatedGroups(
        items=current_page_items,
        page=page,
        per_page=per_page,
        total=total
    )
    
    return render_template(
        'index.html',
        scenario_groups=pagination.items,
        pagination=pagination,
        search=search,
        filter_type=filter_type,
        total_scenarios=len(scenario_groups),
        action_counts=action_counts,
        related_scenarios_count=related_scenarios_count,
        system_settings=system_settings,
        responsible_persons=responsible_persons
    )

def log_auth_event(username, action, status, ip_address, user_agent, details=None):
    """Log authentication events to the database"""
    auth_log = AuthLog(
        username=username,
        action=action,
        status=status,
        ip_address=ip_address,
        user_agent=user_agent,
        details=details
    )
    db.session.add(auth_log)
    db.session.commit()

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    # Get system settings at the start of the function
    settings = SystemSettings.query.first()
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        remember = request.form.get('remember', False)
        
        # First try local authentication
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            # Update last login
            user.last_login = turkey_now()
            db.session.commit()

            # Log successful login
            log_auth_event(
                username=username,
                action='login',
                status='success',
                ip_address=request.remote_addr,
                user_agent=request.user_agent.string,
                details='Local login successful'
            )

            # Log in the user
            login_user(user, remember=remember)

            # Redirect to the user's default landing page or index
            return redirect(url_for(user.default_landing_page or 'index'))
        
        # If local authentication fails and AD is enabled, try AD authentication
        if settings and settings.ad_enabled:
            try:
                print(f"[DEBUG] Attempting AD authentication for user: {username}")
                print(f"[DEBUG] AD Server: {settings.ad_server}")
                print(f"[DEBUG] AD Domain: {settings.ad_domain}")
                print(f"[DEBUG] AD Base DN: {settings.ad_base_dn}")
                print(f"[DEBUG] AD Admin Group: {settings.ad_admin_group}")
                
                # Validate AD server address
                if not settings.ad_server:
                    raise Exception("AD sunucu adresi yapılandırılmamış")
                
                # Connect to AD server
                try:
                    server = ldap3.Server(
                        settings.ad_server,
                        port=settings.ad_port,
                        use_ssl=settings.ad_use_ssl,
                        get_info=ldap3.ALL
                    )
                except Exception as e:
                    print(f"[ERROR] Failed to create AD server object: {str(e)}")
                    raise Exception(f"AD sunucu adresi geçersiz: {settings.ad_server}")
                
                # First try to bind with service account
                try:
                    print(f"[DEBUG] Attempting service account bind with: {settings.ad_username}")
                    conn = ldap3.Connection(
                        server,
                        user=f"{settings.ad_domain}\\{settings.ad_username}",
                        password=settings.ad_password,
                        authentication=ldap3.NTLM,
                        auto_bind=True
                    )
                    print("[DEBUG] Service account bind successful")
                except Exception as e:
                    print(f"[ERROR] Service account bind failed: {str(e)}")
                    # If service account bind fails, try direct user authentication
                    print("[DEBUG] Trying direct user authentication")
                    try:
                        conn = ldap3.Connection(
                            server,
                            user=f"{settings.ad_domain}\\{username}",
                            password=password,
                            authentication=ldap3.NTLM,
                            auto_bind=True
                        )
                        print("[DEBUG] Direct user authentication successful")
                    except Exception as auth_error:
                        print(f"[ERROR] Direct user authentication failed: {str(auth_error)}")
                        raise Exception("Kullanıcı adı veya şifre hatalı")
                
                # Search for user
                search_filter = settings.ad_search_filter.replace('{username}', username)
                print(f"[DEBUG] Searching with filter: {search_filter}")
                print(f"[DEBUG] Search base: {settings.ad_base_dn}")
                
                # Try different search filters if the first one fails
                search_filters = [
                    search_filter,  # Original filter
                    f"(sAMAccountName={username})",  # Simple sAMAccountName filter
                    f"(userPrincipalName={username}@{settings.ad_domain})",  # UPN filter
                    f"(&(objectClass=user)(|(sAMAccountName={username})(userPrincipalName={username}@{settings.ad_domain})))"  # Combined filter
                ]
                
                user_found = False
                user_dn = None
                for current_filter in search_filters:
                    print(f"[DEBUG] Trying filter: {current_filter}")
                    conn.search(
                        search_base=settings.ad_base_dn,
                        search_filter=current_filter,
                        attributes=['memberOf', 'sAMAccountName', 'userPrincipalName', 'displayName', 'mail'],
                        search_scope=ldap3.SUBTREE
                    )
                    print(f"[DEBUG] Search result count: {len(conn.entries)}")
                    if len(conn.entries) > 0:
                        user_found = True
                        user_dn = conn.entries[0].entry_dn
                        print(f"[DEBUG] User found with filter: {current_filter}")
                        print(f"[DEBUG] User DN: {user_dn}")
                        break
                
                if not user_found:
                    print(f"[ERROR] No user found with any filter for username: {username}")
                    raise Exception("Kullanıcı bulunamadı. Lütfen kullanıcı adınızı kontrol edin.")
                
                # Check if user is in admin group
                is_admin = False
                is_authorized = False
                user_full_name = None
                if settings.ad_admin_group:
                    admin_group_dn = settings.ad_admin_group
                    print(f"[DEBUG] Checking admin group membership: {admin_group_dn}")
                    
                    # Search for user's group memberships
                    conn.search(
                        search_base=settings.ad_base_dn,
                        search_filter=f"(&(objectClass=user)(sAMAccountName={username}))",
                        attributes=['memberOf', 'displayName']
                    )
                    print(f"[DEBUG] Group search result count: {len(conn.entries)}")
                    if len(conn.entries) > 0:
                        user_groups = conn.entries[0].memberOf.values
                        user_full_name = conn.entries[0].displayName.value if hasattr(conn.entries[0], 'displayName') else None
                        print(f"[DEBUG] User groups: {user_groups}")
                        print(f"[DEBUG] User full name: {user_full_name}")
                        is_admin = any(admin_group_dn.lower() in group.lower() for group in user_groups)
                        print(f"[DEBUG] Is admin: {is_admin}")
                        # Check if user is member of E152_mft-webui group
                        required_group = "CN=E152_mft-webui,OU=Groups,OU=_GlobalResources,OU=E152,DC=emea,DC=corpdir,DC=net"
                        is_authorized = any(required_group.lower() in group.lower() for group in user_groups)
                        print(f"[DEBUG] Is authorized: {is_authorized}")
                    else:
                        print(f"[ERROR] No group memberships found for user: {username}")
                
                if not is_authorized:
                    error_msg = "Bu uygulamaya erişim yetkiniz bulunmamaktadır. Lütfen sistem yöneticinize başvurun."
                    print(f"[ERROR] {error_msg}")
                    flash(error_msg, 'error')
                    log_auth_event(username, 'login', 'failed', request.remote_addr, request.user_agent.string, error_msg)
                    return render_template('login.html', 
                                        system_settings=settings,
                                        close_window=True,
                                        error_message=error_msg)

                # If we get here, authentication was successful
                # Check if user exists in database
                user = User.query.filter_by(username=username).first()
                if not user:
                    # Create new user
                    user = User(
                        username=username,
                        email=f"{username}@{settings.ad_domain}",
                        is_ad_user=True,
                        is_admin=is_admin,
                        full_name=user_full_name
                    )
                    db.session.add(user)
                    db.session.commit()

                # Update last login
                user.last_login = turkey_now()
                db.session.commit()

                # Log successful login
                log_auth_event(
                    username=username,
                    action='login',
                    status='success',
                    ip_address=request.remote_addr,
                    user_agent=request.user_agent.string,
                    details='AD login successful'
                )

                # Log in the user
                login_user(user, remember=remember)

                # Redirect to the user's default landing page or index
                return redirect(url_for(user.default_landing_page or 'index'))

            except Exception as e:
                print(f"[ERROR] AD authentication error: {str(e)}")
                flash(str(e), 'error')
                log_auth_event(
                    username=username,
                    action='login',
                    status='failed',
                    ip_address=request.remote_addr,
                    user_agent=request.user_agent.string,
                    details=f'AD authentication failed: {str(e)}'
                )
                return render_template('login.html', system_settings=settings)

        # If both local and AD authentication failed
        flash('Geçersiz kullanıcı adı veya şifre.', 'error')
        log_auth_event(
            username=username or 'unknown',
            action='login',
            status='failed',
            ip_address=request.remote_addr,
            user_agent=request.user_agent.string,
            details='Invalid username or password'
        )

    return render_template('login.html', system_settings=settings)

@app.route('/logout')
@login_required
def logout():
    # Log logout event
    auth_log = AuthLog(
        username=current_user.username,
        action='logout',
        status='success',
        ip_address=request.remote_addr,
        user_agent=request.user_agent.string,
        details='User logged out'
    )
    db.session.add(auth_log)
    db.session.commit()
    
    logout_user()
    return redirect(url_for('login'))

@app.route('/scenario/add', methods=['GET', 'POST'])
@login_required
def add_scenario():
    if request.method == 'POST':
        try:
            scenario = Scenario(
                scenario_name=request.form['scenario_name'],
                type_of_action=request.form['type_of_action'],
                user=request.form['user'],
                recurrence=request.form['recurrence'],
                repeat_rate=request.form.get('repeat_rate'),
                source=request.form['source'],
                destination=request.form['destination'],
                email_notification=request.form.get('email_notification'),
                responsible_person=request.form['responsible_person'],
                created_by=current_user.id,
                department=request.form.get('department')
            )

            db.session.add(scenario)
            db.session.commit()

            flash('Scenario has been added successfully.', 'success')
            return redirect(url_for('index'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding scenario: {str(e)}', 'error')
            return redirect(url_for('add_scenario'))

    return render_template('add_scenario.html')


@app.route('/request_change/<int:scenario_id>', methods=['POST'])
@login_required
def request_change(scenario_id):
    try:
        scenario = Scenario.query.get_or_404(scenario_id)
        field_to_change = request.form.get('field_to_change')
        requested_value = request.form.get('requested_value')
        description = request.form.get('description')

        if not field_to_change or not requested_value or not description:
            flash('All fields are required.', 'error')
            return redirect(url_for('scenario_detail', scenario_id=scenario_id))

        current_value = getattr(scenario, field_to_change)

        change_request = ChangeRequest(
            scenario_id=scenario_id,
            requested_by=current_user.id,
            field_to_change=field_to_change,
            current_value=str(current_value),
            requested_value=requested_value,
            description=description
        )

        db.session.add(change_request)
        db.session.commit()

        flash('Change request has been submitted successfully.', 'success')
        return redirect(url_for('scenario_detail', scenario_id=scenario_id))
    except Exception as e:
        db.session.rollback()
        flash(f'Error submitting change request: {str(e)}', 'error')
        return redirect(url_for('scenario_detail', scenario_id=scenario_id))


@app.route('/import-scenarios', methods=['GET', 'POST'])
@login_required
def import_scenarios():
    # Eğer dönüştürülmüş Excel dosyası varsa, onu otomatik olarak seç
    converted_file = request.args.get('converted_file')
    if converted_file:
        # Geçici dizin oluştur
        temp_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'temp')
        os.makedirs(temp_dir, exist_ok=True)

        # Dönüştürülmüş dosyanın tam yolu
        converted_file_path = os.path.join(
            app.config['UPLOAD_FOLDER'], converted_file)

        # Dosyanın varlığını kontrol et
        if os.path.exists(converted_file_path):
            # Dosyayı geçici dizine kopyala
            temp_file = os.path.join(temp_dir, converted_file)
            import shutil
            shutil.copy2(converted_file_path, temp_file)

            # Geçici dosya yolunu session'da sakla
            session['temp_file_path'] = temp_file

            # Dosyayı oku ve önizleme verilerini hazırla
            try:
                df = pd.read_excel(temp_file)

                # En yüksek Custom ID'yi bul ve 100'ü çıkar
                max_custom_id = 0
                for custom_id in df['Custom ID']:
                    try:
                        custom_id_num = int(custom_id)
                        max_custom_id = max(max_custom_id, custom_id_num)
                    except (ValueError, TypeError):
                        continue

                total_scenarios = max_custom_id - 100 if max_custom_id > 100 else 0

                # Custom ID'ye göre grupla ve her grup için geçerlilik kontrolü
                # yap
                active_count = 0
                passive_count = 0

                for custom_id in df['Custom ID'].unique():
                    if pd.isna(custom_id):
                        continue

                    group = df[df['Custom ID'] == custom_id]

                    # Gruptaki her satır için zorunlu alanları kontrol et
                    is_valid = True
                    for _, row in group.iterrows():
                        if not all(str(row[field]).strip() for field in [
                                   'Scenario Name', 'Recurrence', 'Action Type', 'Source', 'Destination']):
                            is_valid = False
                            break

                    # Status kontrolü
                    status = str(group.iloc[0].get('Status', '')).lower().strip()
                    if status == 'aktif':
                        status = 'active'
                    elif status == 'pasif':
                        status = 'passive'
                    else:
                        status = 'active'  # Varsayılan olarak aktif

                    if is_valid:
                        if status == 'active':
                            active_count += 1
                        else:
                            passive_count += 1

                # Önizleme için ilk 10 senaryoyu hazırla
                preview_data = []
                for _, row in df.head(10).iterrows():
                    # Status dönüşümü
                    status = str(row.get('Status', '')).lower().strip()
                    if status == 'aktif':
                        status = 'active'
                    elif status == 'pasif':
                        status = 'passive'
                    else:
                        status = 'active'  # Varsayılan olarak aktif

                    preview_data.append({
                        'custom_id': str(row.get('Custom ID', '')),
                        'scenario_name': str(row.get('Scenario Name', '')),
                        'recurrence': str(row.get('Recurrence', '')),
                        'repeat_rate': str(row.get('Repeat Rate', '')),
                        'type_of_action': str(row.get('Action Type', '')),
                        'type': str(row.get('Type', '')),
                        'step_count': str(row.get('Step Count', '')),
                        'source': str(row.get('Source', '')),
                        'destination': str(row.get('Destination', '')),
                        'status': status,
                        'folder_id': str(row.get('Folder ID', '')),
                        'department': str(row.get('Department', ''))
                    })

                return render_template('import-scenarios.html',
                                       converted_file=converted_file,
                                       preview_data={
                                           'success': True,
                                           'total': total_scenarios,
                                           'active': active_count,
                                           'passive': passive_count,
                                           'preview': preview_data
                                       })

            except Exception as e:
                # Hata durumunda geçici dosyayı temizle
                if os.path.exists(temp_file):
                    try:
                        os.remove(temp_file)
                    except BaseException:
                        pass
                session.pop('temp_file_path', None)
                flash(f'Dosya okuma hatası: {str(e)}', 'error')
                return render_template('import-scenarios.html')

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Dosya seçilmedi', 'error')
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash('Dosya seçilmedi', 'error')
            return redirect(request.url)

        if file and file.filename.endswith(('.xlsx', '.csv')):
            try:
                # Excel dosyasını oku
                df = pd.read_excel(file) if file.filename.endswith(
                    '.xlsx') else pd.read_csv(file)

                # Gerekli alanları kontrol et
                required_fields = [
                    'Scenario Name',
                    'Recurrence',
                    'Action Type',
                    'Source',
                    'Destination']
                missing_fields = [
                    field for field in required_fields if field not in df.columns]

                if missing_fields:
                    flash(
                        f'Eksik alanlar: {", ".join(missing_fields)}',
                        'error')
                    return redirect(request.url)

                # Verileri doğrula ve dönüştür
                scenarios = []
                for _, row in df.iterrows():
                    scenario = Scenario(
                        scenario_name=row['Scenario Name'],
                        type_of_action=row['Action Type'],
                        user=current_user.username,
                        recurrence=row['Recurrence'],
                        source=row['Source'],
                        destination=row['Destination'],
                        email_notification=row.get('Email Notification'),
                        responsible_person=current_user.username,
                        created_by=current_user.id,
                        status=row.get('Status', 'active'),
                        folder_id=row.get('Folder ID'),
                        type=row.get('Type'),
                        step_count=row.get('Step Count')
                    )
                    scenarios.append(scenario)

                # Veritabanına kaydet
                db.session.bulk_save_objects(scenarios)
                db.session.commit()

                flash(
                    f'{len(scenarios)} senaryo başarıyla içe aktarıldı',
                    'success')
                return redirect(url_for('index'))

            except Exception as e:
                db.session.rollback()
                flash(f'Hata oluştu: {str(e)}', 'error')
                return redirect(request.url)
        else:
            flash('Desteklenmeyen dosya formatı', 'error')
            return redirect(request.url)

    return render_template('import-scenarios.html')


@app.route('/export-scenarios')
@login_required
def export_scenarios():
    # Check if user has permission to export
    if not current_user.is_admin and 'excel_export' not in [role.name for role in current_user.roles]:
        flash('Bu işlem için yetkiniz yok.', 'error')
        return redirect(url_for('unauthorized'))

    # Get all scenarios
    scenarios = Scenario.query.all()
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Senaryolar"
    
    # Add headers
    headers = ['Senaryo Adı', 'İşlem Tipi', 'Kullanıcı', 'Tekrar', 'Tekrar Oranı', 
               'Kaynak', 'Hedef', 'E-posta Bildirimi', 'Sorumlu Kişi', 'Durum', 
               'Oluşturulma Tarihi', 'Grup ID', 'Klasör ID', 'Tip', 'Adım Sayısı', 'Departman']
    ws.append(headers)
    
    # Add data
    for scenario in scenarios:
        ws.append([
            scenario.scenario_name,
            scenario.type_of_action,
            scenario.user,
            scenario.recurrence,
            scenario.repeat_rate,
            scenario.source,
            scenario.destination,
            scenario.email_notification,
            scenario.responsible_person,
            scenario.status,
            scenario.created_at.strftime('%d.%m.%Y %H:%M') if scenario.created_at else '',
            scenario.group_id,
            scenario.folder_id,
            scenario.type,
            scenario.step_count,
            scenario.department
        ])
    
    # Create a temporary file
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_file.close()
    
    # Save the workbook to the temporary file
    wb.save(temp_file.name)
    
    # Send the file
    response = send_file(
        temp_file.name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='senaryolar.xlsx'
    )
    
    # Clean up the temporary file after sending
    @response.call_on_close
    def cleanup():
        try:
            os.unlink(temp_file.name)
        except:
            pass
    
    return response


def get_base_scenario_name(scenario_name):
    """Senaryo adından temel ismi çıkarır"""
    # MB numarasına göre base_name oluştur
    mb_match = re.search(r'(.*MB-\d+)', scenario_name)
    if mb_match:
        return mb_match.group(1)

    # MB numarası yoksa DOWNLOAD/UPLOAD/AZURE-UPLOAD/TO-AZURE kısımlarını
    # kaldır
    return re.sub(
        r'_(?:DOWNLOAD|UPLOAD|AZURE-UPLOAD|TO-AZURE)(?:\d+)?$',
        '',
        scenario_name)


def check_step_count_consistency(scenario_name):
    """Bir senaryo için ilişkili tüm senaryoları bulur ve sayısını döndürür"""
    base_name = get_base_scenario_name(scenario_name)
    all_related = Scenario.query.filter(
        Scenario.scenario_name.like(f"{base_name}%")
    ).all()
    return len(all_related)


@app.route('/scenario/<int:scenario_id>')
@login_required
def scenario_detail(scenario_id):
    # Kullanıcının admin olup olmadığını kontrol et
    is_admin = current_user.is_admin

    # Kullanıcının scenario_viewer rolüne sahip olup olmadığını kontrol et
    has_scenario_viewer_role = False
    for role in current_user.roles:
        if role.name == 'scenario_viewer':
            has_scenario_viewer_role = True
            break

    # Eğer kullanıcı admin değilse ve scenario_viewer rolüne sahip değilse
    # erişimi reddet
    if not is_admin and not has_scenario_viewer_role:
        flash(
            'Bu sayfayı görüntülemek için gerekli yetkiye sahip değilsiniz.',
            'error')
        return render_template('access_denied.html')

    # Get the specific scenario by ID
    scenario = Scenario.query.get_or_404(scenario_id)
    print(f"[DEBUG] Scenario detail - responsible_person: {scenario.responsible_person}")

    # Get base scenario name
    base_name = get_base_scenario_name(scenario.scenario_name)

    # Get all related scenarios
    related_scenarios = Scenario.query.filter(
        Scenario.scenario_name.like(f"{base_name}%")
    ).order_by(Scenario.scenario_name).all()

    # Get the step count from index page logic
    index_step_count = check_step_count_consistency(scenario.scenario_name)
    actual_step_count = len(related_scenarios)
    has_step_count_mismatch = index_step_count != actual_step_count

    # Double check by querying again
    scenario_check = Scenario.query.get(scenario_id)
    print(f"[DEBUG] Scenario detail (re-query) - responsible_person: {scenario_check.responsible_person}")

    return render_template('scenario_detail.html',
                           scenario=scenario,
                           related_scenarios=related_scenarios,
                           has_step_count_mismatch=has_step_count_mismatch,
                           index_step_count=index_step_count)


@app.route('/scenario/edit/<int:scenario_id>', methods=['GET', 'POST'])
@login_required
def edit_scenario(scenario_id):
    scenario = Scenario.query.get_or_404(scenario_id)
    if request.method == 'POST':
        # Get the responsible person from the form
        responsible_person = request.form.get('responsible_person')
        
        # Update all fields
        scenario.scenario_name = request.form.get('scenario_name')
        scenario.type_of_action = request.form.get('type_of_action')
        scenario.user = request.form.get('user')
        scenario.recurrence = request.form.get('recurrence')
        scenario.source = request.form.get('source')
        scenario.destination = request.form.get('destination')
        scenario.email_notification = request.form.get('email_notification')
        scenario.responsible_person = responsible_person

        print(f"[DEBUG] Before commit - scenario.responsible_person: {scenario.responsible_person}")
        try:
            db.session.commit()
            print(f"[DEBUG] After commit - scenario.responsible_person: {scenario.responsible_person}")
            # Verify the update by reloading from database
            db.session.refresh(scenario)
            print(f"[DEBUG] After refresh - scenario.responsible_person: {scenario.responsible_person}")
            # Double check by querying again
            updated_scenario = Scenario.query.get(scenario_id)
            print(f"[DEBUG] After re-query - responsible_person: {updated_scenario.responsible_person}")
            flash('Scenario updated successfully')
            return redirect(url_for('scenario_detail', scenario_id=scenario.id))
        except Exception as e:
            print(f"[DEBUG] Error updating scenario: {str(e)}")
            db.session.rollback()
            flash('Error updating scenario', 'error')
            return redirect(url_for('scenario_detail', scenario_id=scenario.id))

    return render_template('edit_scenario.html', scenario=scenario)


@app.route('/change-requests')
@login_required
def change_requests():
    if not current_user.is_admin:
        flash('Only administrators can view change requests.', 'warning')
        return redirect(url_for('index'))

    change_requests = ChangeRequest.query.order_by(
        ChangeRequest.created_at.desc()).all()
    return render_template(
        'change_requests.html',
        change_requests=change_requests)


@app.route('/approve-change/<int:request_id>', methods=['POST'])
@login_required
def approve_change(request_id):
    if not current_user.is_admin:
        flash('Only administrators can approve change requests.', 'warning')
        return redirect(url_for('index'))

    try:
        change_request = ChangeRequest.query.get_or_404(request_id)
        if change_request.status != 'pending':
            flash('This change request has already been processed.', 'warning')
            return redirect(url_for('change_requests'))

        scenario = Scenario.query.get(change_request.scenario_id)

        # Update the scenario with the requested change
        setattr(
            scenario,
            change_request.field_to_change,
            change_request.requested_value)

        # Update change request status
        change_request.status = 'approved'

        db.session.commit()
        flash('Change request has been approved and applied.', 'success')
        return redirect(url_for('change_requests'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error approving change request: {str(e)}', 'error')
        return redirect(url_for('change_requests'))


@app.route('/reject-change/<int:request_id>', methods=['POST'])
@login_required
def reject_change(request_id):
    if not current_user.is_admin:
        flash('Only administrators can reject change requests.', 'warning')
        return redirect(url_for('index'))

    try:
        change_request = ChangeRequest.query.get_or_404(request_id)
        if change_request.status != 'pending':
            flash('This change request has already been processed.', 'warning')
            return redirect(url_for('change_requests'))

        change_request.status = 'rejected'
        db.session.commit()

        flash('Change request has been rejected.', 'warning')
        return redirect(url_for('change_requests'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error rejecting change request: {str(e)}', 'error')
        return redirect(url_for('change_requests'))

# A simple form class for CSRF protection


class EmptyForm(FlaskForm):
    pass

# Add context processor to make form available in all templates


@app.context_processor
def utility_processor():
    def get_empty_form():
        return EmptyForm()

    def get_pending_requests():
        if current_user.is_authenticated and current_user.is_admin:
            return ChangeRequest.query.filter_by(status='pending').count()
        return 0

    def get_system_settings():
        settings = SystemSettings.query.first()
        if not settings:
            settings = SystemSettings()
            db.session.add(settings)
            db.session.commit()
        return settings

    def get_pending_changes_count():
        if not current_user.is_authenticated:
            return 0

        if current_user.is_admin:
            return ChangeRequest.query.filter_by(status='pending').count()
        return 0

    return dict(
        get_pending_requests=get_pending_requests,
        get_system_settings=get_system_settings,
        get_pending_changes_count=get_pending_changes_count,
        get_empty_form=get_empty_form
    )


@app.route('/system_settings', methods=['GET', 'POST'])
@login_required
@csrf.exempt
def system_settings():
    # Kullanıcının admin olup olmadığını kontrol et
    if not current_user.is_admin:
        flash('Bu sayfaya erişim yetkiniz yok.', 'error')
        return redirect(url_for('index'))

    settings = SystemSettings.query.first()
    if not settings:
        settings = SystemSettings()
        db.session.add(settings)
        db.session.commit()

    if request.method == 'POST':
        settings.system_name = request.form.get('system_name')
        settings.timezone = request.form.get('timezone')
        settings.date_format = request.form.get('date_format')
        settings.min_password_length = int(
            request.form.get('min_password_length', 8))
        settings.session_timeout = int(request.form.get('session_timeout', 30))
        settings.two_factor = request.form.get('two_factor') == 'on'
        settings.backup_frequency = request.form.get('backup_frequency')
        settings.backup_location = request.form.get('backup_location')
        settings.backup_retention = int(
            request.form.get('backup_retention', 30))
        settings.log_level = request.form.get('log_level')
        settings.log_path = request.form.get('log_path')
        settings.log_rotation = request.form.get('log_rotation') == 'on'
        settings.login_method = request.form.get('login_method') == 'on'
        
        # Active Directory Settings
        settings.ad_enabled = request.form.get('ad_enabled') == 'on'
        settings.ad_server = request.form.get('ad_server')
        settings.ad_domain = request.form.get('ad_domain')
        settings.ad_base_dn = request.form.get('ad_base_dn')
        settings.ad_username = request.form.get('ad_username')
        settings.ad_password = request.form.get('ad_password')
        settings.ad_port = int(request.form.get('ad_port', 389))
        settings.ad_use_ssl = request.form.get('ad_use_ssl') == 'on'
        settings.ad_search_filter = request.form.get('ad_search_filter', '(&(objectClass=user)(sAMAccountName={username})(memberOf=CN=E152_mft-webui,OU=Groups,OU=_GlobalResources,OU=E152,DC=emea,DC=corpdir,DC=net))')
        settings.ad_group_filter = request.form.get('ad_group_filter', '(memberOf={group})')
        settings.ad_admin_group = request.form.get('ad_admin_group', 'CN=E152_mft-webui,OU=Groups,OU=_GlobalResources,OU=E152,DC=emea,DC=corpdir,DC=net')

        db.session.commit()
        flash('Sistem ayarları başarıyla güncellendi.', 'success')
        return redirect(url_for('system_settings'))

    return render_template('system_settings.html', settings=settings)


@app.route('/user_management')
@login_required
def user_management():
    # Kullanıcı yönetimi için yetki kontrolü
    has_user_manager_role = any(
        role.name == 'user_manager' for role in current_user.roles)
    if not current_user.is_admin and not has_user_manager_role:
        flash('Bu sayfaya erişim izniniz yok.', 'error')
        return redirect(url_for('unauthorized'))

    users = User.query.all()
    ms_users_count = User.query.filter_by(is_oauth_user=True).count()
    admin_count = User.query.filter_by(is_admin=True).count()
    total_users = len(users)
    allowed_emails = AllowedEmail.query.all()

    # Get all roles for the dropdown
    roles = Role.query.all()

    return render_template('user_management.html',
                           users=users,
                           total_users=total_users,
                           admin_count=admin_count,
                           ms_users_count=ms_users_count,
                           allowed_emails=allowed_emails,
                           roles=roles)


@app.route('/add_user', methods=['POST'])
@login_required
def add_user():
    has_user_manager_role = any(
        role.name == 'user_manager' for role in current_user.roles)
    if not current_user.is_admin and not has_user_manager_role:
        return jsonify({'success': False,
                        'message': 'Bu işlem için yetkiniz yok.'})

    auth_method = request.form.get('auth_method')
    email = request.form.get('email')
    full_name = request.form.get('full_name')

    try:
        if auth_method == 'local':
            username = request.form.get('username')
            password = request.form.get('password')
            confirm_password = request.form.get('confirm_password')

            # Check if username already exists
            if User.query.filter_by(username=username).first():
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify(
                        {'success': False, 'message': 'Bu kullanıcı adı zaten kullanılıyor.'})
                flash('Bu kullanıcı adı zaten kullanılıyor.', 'error')
                return redirect(url_for('user_management'))

            if password != confirm_password:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify(
                        {'success': False, 'message': 'Şifreler eşleşmiyor.'})
                flash('Şifreler eşleşmiyor.', 'error')
                return redirect(url_for('user_management'))

            user = User(
                username=username,
                email=email,
                password_hash=generate_password_hash(password),
                full_name=full_name,
                is_oauth_user=False
            )
        else:  # Microsoft authentication
            # For Microsoft users, generate a random username based on email
            username = email.split('@')[0]
            # Add a random string to ensure uniqueness
            while User.query.filter_by(username=username).first():
                username = f"{email.split('@')[0]}_{secrets.token_hex(4)}"

            user = User(
                username=username,
                email=email,
                full_name=full_name,
                is_oauth_user=True
            )

        db.session.add(user)
        
        # Log user creation
        auth_log = AuthLog(
            username=current_user.username,
            action='user_created',
            status='success',
            ip_address=request.remote_addr,
            details=f'Yeni kullanıcı oluşturuldu: {username} ({auth_method})',
            target_user=username,
            performed_by=current_user.username
        )
        db.session.add(auth_log)
        
        db.session.commit()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify(
                {'success': True, 'message': 'Kullanıcı başarıyla eklendi. Rol atamak için Rol Yönetimi sayfasını kullanın.'})

        flash(
            'Kullanıcı başarıyla eklendi. Rol atamak için Rol Yönetimi sayfasını kullanın.',
            'success')
        return redirect(url_for('user_management'))

    except Exception as e:
        db.session.rollback()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify(
                {'success': False, 'message': f'Kullanıcı eklenirken bir hata oluştu: {str(e)}'})
        flash(f'Kullanıcı eklenirken bir hata oluştu: {str(e)}', 'error')
        return redirect(url_for('user_management'))


@app.route('/edit_user/<int:user_id>', methods=['POST'])
@login_required
def edit_user(user_id):
    has_user_manager_role = any(
        role.name == 'user_manager' for role in current_user.roles)
    if not current_user.is_admin and not has_user_manager_role:
        return {'success': False, 'message': 'Bu işlem için yetkiniz yok.'}, 403

    user = User.query.get_or_404(user_id)

    # Prevent editing the admin user
    if user.username == 'admin':
        return {'success': False,
                'message': 'Sistem yöneticisi (admin) kullanıcısı düzenlenemez.'}, 403

    username = request.form.get('username')
    email = request.form.get('email')
    full_name = request.form.get('full_name')
    password = request.form.get('password')
    is_admin = request.form.get('is_admin') == 'on'

    existing_user = User.query.filter_by(username=username).first()
    if existing_user and existing_user.id != user_id:
        return {'success': False,
                'message': 'Bu kullanıcı adı zaten kullanılıyor.'}, 400

    try:
        user.username = username
        user.email = email
        user.full_name = full_name
        if password:
            user.password_hash = generate_password_hash(password)
        user.is_admin = is_admin

        db.session.commit()
        return {'success': True,
                'message': 'Kullanıcı başarıyla güncellendi.'}, 200
    except Exception as e:
        db.session.rollback()
        return {'success': False,
                'message': 'Kullanıcı güncellenirken bir hata oluştu.'}, 500


@app.route('/delete_user/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    has_user_manager_role = any(
        role.name == 'user_manager' for role in current_user.roles)
    if not current_user.is_admin and not has_user_manager_role:
        return jsonify(
            {'success': False, 'message': 'Bu işlem için yetkiniz yok.'}), 403

    user = User.query.get_or_404(user_id)

    # Prevent deleting the admin user
    if user.username == 'admin':
        return jsonify(
            {'success': False, 'message': 'Sistem yöneticisi (admin) kullanıcısı silinemez.'}), 403

    if user.is_admin:
        return jsonify(
            {'success': False, 'message': 'Admin kullanıcısı silinemez.'}), 400

    try:
        # Log user deletion before actually deleting
        auth_log = AuthLog(
            username=current_user.username,
            action='user_deleted',
            status='success',
            ip_address=request.remote_addr,
            details=f'Kullanıcı silindi: {user.username}',
            target_user=user.username,
            performed_by=current_user.username
        )
        db.session.add(auth_log)

        # Handle foreign key relationships
        # 1. Remove user roles
        UserRole.query.filter_by(user_id=user.id).delete()
        
        # 2. Update scenarios created by this user
        Scenario.query.filter_by(created_by=user.id).update({Scenario.created_by: None})
        
        # 3. Update change requests
        ChangeRequest.query.filter_by(requested_by=user.id).update({ChangeRequest.requested_by: None})
        
        # 4. Update logs
        Log.query.filter_by(created_by=user.id).update({Log.created_by: None})
        
        # 5. Update allowed emails
        AllowedEmail.query.filter_by(added_by=user.id).update({AllowedEmail.added_by: None})
        
        # Now delete the user
        db.session.delete(user)
        db.session.commit()
        
        return jsonify({'success': True,
                        'message': 'Kullanıcı başarıyla silindi.'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting user: {str(e)}")
        return jsonify(
            {'success': False, 'message': f'Kullanıcı silinirken bir hata oluştu: {str(e)}'}), 500


@app.route('/email_settings')
@login_required
def email_settings():
    if not current_user.is_admin:
        flash('Bu sayfaya erişim yetkiniz yok.', 'error')
        return redirect(url_for('index'))
    return render_template('email_settings.html')


@app.route('/login/microsoft')
@csrf.exempt
def microsoft_login():
    # Generate a random state value for security
    state = secrets.token_hex(16)
    session['state'] = state

    # Generate the Microsoft login URL
    auth_url = msal_app.get_authorization_request_url(
        scopes=AZURE_SCOPES,
        redirect_uri=REDIRECT_URI,
        state=state,
        response_type="code"
    )
    return redirect(auth_url)


@app.route('/login/microsoft/callback')
@csrf.exempt
def microsoft_login_callback():
    try:
        # Get the authorization code from the callback
        code = request.args.get('code')
        if not code:
            flash('Microsoft girişi başarısız oldu.', 'error')
            return redirect(url_for('login'))

        # Exchange the code for an access token
        token_response = requests.post(
            'https://login.microsoftonline.com/common/oauth2/v2.0/token',
            data={
                'client_id': app.config['MICROSOFT_CLIENT_ID'],
                'client_secret': app.config['MICROSOFT_CLIENT_SECRET'],
                'code': code,
                'redirect_uri': url_for('microsoft_login_callback', _external=True),
                'grant_type': 'authorization_code'
            }
        )

        if token_response.status_code != 200:
            flash('Microsoft token alınamadı.', 'error')
            return redirect(url_for('login'))

        token_data = token_response.json()
        access_token = token_data['access_token']

        # Get user info from Microsoft Graph API
        user_response = requests.get(
            'https://graph.microsoft.com/v1.0/me',
            headers={'Authorization': f'Bearer {access_token}'}
        )

        if user_response.status_code != 200:
            flash('Microsoft kullanıcı bilgileri alınamadı.', 'error')
            return redirect(url_for('login'))

        user_data = user_response.json()
        email = user_data.get('mail') or user_data.get('userPrincipalName')
        ms_id = user_data.get('id')
        display_name = user_data.get('displayName')

        if not email:
            flash('Microsoft hesabınızdan e-posta adresi alınamadı.', 'error')
            return redirect(url_for('login'))

        # Check if user exists
        user = User.query.filter_by(email=email).first()

        if not user:
            # Create new user
            user = User(
                username=email.split('@')[0],  # Use part before @ as username
                email=email,
                ms_id=ms_id,
                is_oauth_user=True,
                is_ad_user=True,
                full_name=display_name
            )
            db.session.add(user)
            db.session.commit()

            # Log the new user creation
            log_auth_event(
                username=user.username,
                action='user_created',
                status='success',
                ip_address=request.remote_addr,
                user_agent=request.user_agent.string,
                details=f'User created via Microsoft login: {email}'
            )

        # Update last login
        user.last_login = turkey_now()
        db.session.commit()

        # Log successful login
        log_auth_event(
            username=user.username,
            action='login',
            status='success',
            ip_address=request.remote_addr,
            user_agent=request.user_agent.string,
            details='Microsoft login successful'
        )

        # Log in the user
        login_user(user)

        # Redirect to the user's default landing page or index
        return redirect(url_for(user.default_landing_page or 'index'))

    except Exception as e:
        print(f"Microsoft login error: {str(e)}")
        flash('Microsoft girişi sırasında bir hata oluştu.', 'error')
        return redirect(url_for('login'))


@app.route('/allowed_emails')
@login_required
def allowed_emails():
    if not current_user.is_admin:
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('index'))

    # Veritabanından izinli e-posta adreslerini al
    allowed_emails_list = AllowedEmail.query.all()

    # `added_by` değeri None olanlar için mevcut kullanıcıyı ata
    for email in allowed_emails_list:
        if email.added_by is None:
            email.added_by = current_user.id

    # Değişiklikleri kaydet
    db.session.commit()

    return render_template(
        'allowed_emails.html',
        allowed_emails=allowed_emails_list)


@app.route('/add_allowed_email', methods=['POST'])
@login_required
@csrf.exempt
def add_allowed_email():
    if not current_user.is_admin:
        flash('You do not have permission to perform this action.', 'error')
        return redirect(url_for('index'))

    email = request.form.get('email')
    # Get admin status from form
    is_admin = request.form.get('is_admin') == 'on'

    if not email:
        flash('Email is required.', 'error')
        return redirect(url_for('allowed_emails'))

    existing = AllowedEmail.query.filter_by(email=email).first()
    if existing:
        if existing.is_active:
            flash('This email is already allowed.', 'warning')
        else:
            existing.is_active = True
            existing.is_admin = is_admin  # Update admin status
            db.session.commit()
            flash('Email access has been restored.', 'success')
        return redirect(url_for('allowed_emails'))

    allowed_email = AllowedEmail(
        email=email,
        added_by=current_user.id,
        is_admin=is_admin  # Set admin status
    )
    db.session.add(allowed_email)
    db.session.commit()
    flash('Email has been added to allowed list.', 'success')
    return redirect(url_for('allowed_emails'))


@app.route('/remove_allowed_email/<int:email_id>', methods=['POST'])
@login_required
def remove_allowed_email(email_id):
    if not current_user.is_admin:
        flash('You do not have permission to perform this action.', 'error')
        return redirect(url_for('index'))

    allowed_email = AllowedEmail.query.get_or_404(email_id)
    allowed_email.is_active = False
    db.session.commit()
    flash('Email has been removed from allowed list.', 'success')
    return redirect(url_for('allowed_emails'))


@app.route('/scenario/toggle_status/<int:scenario_id>', methods=['POST'])
@login_required
def toggle_scenario_status(scenario_id):
    # Check if user has admin or scenario_manager role
    if not current_user.is_admin and 'scenario_manager' not in [role.name for role in current_user.roles]:
        flash('Bu işlem için yetkiniz bulunmamaktadır.', 'error')
        return redirect(url_for('index'))

    try:
        scenario = Scenario.query.get_or_404(scenario_id)
        old_status = scenario.status
        new_status = 'inactive' if scenario.status == 'active' else 'active'
        scenario.status = new_status
        
        # Create log entry for status change
        log = Log(
            scenario_id=scenario.id,
            timestamp=turkey_now(),
            log_type='manual',
            level='INFO',
            message=f'Senaryo durumu değiştirildi: {old_status} -> {new_status}',
            source='Web Interface',
            created_by=current_user.id,
            category='status_change',
            host=request.remote_addr,
            process='toggle_scenario_status'
        )
        db.session.add(log)
        
        db.session.commit()
        flash(f'Senaryo durumu başarıyla güncellendi.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Senaryo durumu güncellenirken bir hata oluştu: {str(e)}', 'error')
    
    return redirect(url_for('index'))


@app.route('/import-json', methods=['GET'])
@login_required
def import_json():
    return render_template('import_json.html')


@app.route('/import-json/preview', methods=['POST'])
@login_required
def import_json_preview():
    if 'json_file' not in request.files:
        flash('Lütfen bir JSON dosyası seçin.', 'error')
        return redirect(url_for('import_json'))

    file = request.files['json_file']
    if file.filename == '':
        flash('Dosya seçilmedi.', 'error')
        return redirect(url_for('import_json'))

    if not file.filename.endswith('.json'):
        flash('Lütfen geçerli bir JSON dosyası seçin.', 'error')
        return redirect(url_for('import_json'))

    try:
        # JSON dosyasını oku
        json_data = json.load(file)

        # İlk kaydı örnek veri olarak kullan
        preview_data = json_data[0] if isinstance(
            json_data, list) else json_data

        # Session'a verileri kaydet
        session['json_data'] = json_data
        session['current_field_index'] = 0
        session['field_mappings'] = {}
        session['skipped_fields'] = []

        return render_template('import_json.html',
                               preview_data=preview_data,
                               json_fields=list(preview_data.keys()),
                               target_fields=TARGET_FIELDS,
                               current_field_index=0,
                               total_fields=len(TARGET_FIELDS),
                               field_mappings={},
                               skipped_fields=[])

    except json.JSONDecodeError:
        flash('Geçersiz JSON formatı.', 'error')
        return redirect(url_for('import_json'))
    except Exception as e:
        flash(f'Dosya yüklenirken bir hata oluştu: {str(e)}', 'error')
        return redirect(url_for('import_json'))


@app.route('/import-json/import', methods=['POST'])
@login_required
def import_json_data():
    action = request.form.get('action')
    current_field_index = int(request.form.get('current_field_index', 0))

    # Session'dan verileri al
    json_data = session.get('json_data', [])
    field_mappings = session.get('field_mappings', {})
    skipped_fields = session.get('skipped_fields', [])

    if not json_data:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify(
                {'success': False, 'message': 'Oturum süresi doldu. Lütfen tekrar dosya yükleyin.'})
        flash('Oturum süresi doldu. Lütfen tekrar dosya yükleyin.', 'error')
        return redirect(url_for('import_json'))

    # Mevcut alan için eşleştirme bilgisini kaydet
    current_field = TARGET_FIELDS[current_field_index]
    skip_field = request.form.get('skip_field') == 'on'

    if skip_field:
        if current_field not in skipped_fields:
            skipped_fields.append(current_field)
        if current_field in field_mappings:
            del field_mappings[current_field]
    else:
        source_field = request.form.get('source_field')
        static_value = request.form.get('static_value')

        if static_value:
            # Sabit değer girişi için özel işlem
            try:
                # JSON formatındaki değeri temizle
                static_value = static_value.strip()
                if ':' in static_value:
                    # "Key": "Value" formatından sadece value'yu al
                    static_value = static_value.split(':', 1)[1].strip()
                    # Başındaki ve sonundaki tırnak işaretlerini kaldır
                    static_value = static_value.strip('"\'')
                field_mappings[current_field] = f"__static__{static_value}"
            except Exception as e:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify(
                        {'success': False, 'message': f'Sabit değer işlenirken hata oluştu: {str(e)}'})
                flash(f'Sabit değer işlenirken hata oluştu: {str(e)}', 'error')
        elif source_field:
            field_mappings[current_field] = source_field

        if current_field in skipped_fields:
            skipped_fields.remove(current_field)

    # Session'ı güncelle
    session['field_mappings'] = field_mappings
    session['skipped_fields'] = skipped_fields

    # Önceki/Sonraki adıma git veya içe aktarmayı tamamla
    if action == 'previous':
        current_field_index = max(0, current_field_index - 1)
    elif action == 'next':
        current_field_index = min(
            len(TARGET_FIELDS) - 1,
            current_field_index + 1)
    elif action == 'finish':
        try:
            # JSON verilerini senaryolara dönüştür
            scenarios = []
            json_records = json_data if isinstance(
                json_data, list) else [json_data]

            for record in json_records:
                scenario_data = {}
                for target_field, source_field in field_mappings.items():
                    if target_field not in skipped_fields:
                        # Alan isimlerini veritabanı alan isimlerine dönüştür
                        db_field = convert_field_name(target_field)

                        # Sabit değer kontrolü
                        if source_field.startswith('__static__'):
                            # '__static__' prefix'ini kaldır
                            scenario_data[db_field] = source_field[9:]
                        elif source_field in record:
                            scenario_data[db_field] = record[source_field]

                # Zorunlu alanları kontrol et
                if not all(
                    field in scenario_data for field in [
                        'scenario_name',
                        'recurrence',
                        'type_of_action',
                        'source',
                        'destination']):
                    continue

                # Yeni senaryo oluştur
                scenario = Scenario(
                    scenario_name=scenario_data.get('scenario_name'),
                    type_of_action=scenario_data.get('type_of_action'),
                    recurrence=scenario_data.get(
                        'recurrence',
                        'Daily'),
                    source=scenario_data.get('source'),
                    destination=scenario_data.get('destination'),
                    email_notification=scenario_data.get('email_notification'),
                    responsible_person=scenario_data.get(
                        'responsible_person',
                        current_user.username),
                    created_by=current_user.id,
                    created_at=turkey_now())
                scenarios.append(scenario)

            # Senaryoları veritabanına kaydet
            if scenarios:
                db.session.bulk_save_objects(scenarios)
                db.session.commit()

                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    # AJAX isteği ise JSON yanıtı döndür
                    return jsonify({
                        'success': True,
                        'message': f'{len(scenarios)} senaryo başarıyla içe aktarıldı.',
                        'redirect': url_for('index')
                    })

                flash(
                    f'{len(scenarios)} senaryo başarıyla içe aktarıldı.',
                    'success')
            else:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({
                        'success': False,
                        'message': 'İçe aktarılacak geçerli senaryo bulunamadı.'
                    })
                flash('İçe aktarılacak geçerli senaryo bulunamadı.', 'warning')

            # Session'ı temizle
            session.pop('json_data', None)
            session.pop('field_mappings', None)
            session.pop('skipped_fields', None)
            session.pop('current_field_index', None)

            return redirect(url_for('index'))

        except Exception as e:
            db.session.rollback()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({
                    'success': False,
                    'message': f'Senaryolar içe aktarılırken bir hata oluştu: {str(e)}'
                })
            flash(
                f'Senaryolar içe aktarılırken bir hata oluştu: {str(e)}',
                'error')
            return redirect(url_for('import_json'))

    # Session'ı güncelle
    session['current_field_index'] = current_field_index

    # İlk kaydı örnek veri olarak kullan
    preview_data = json_data[0] if isinstance(json_data, list) else json_data

    return render_template('import_json.html',
                           preview_data=preview_data,
                           json_fields=list(preview_data.keys()),
                           target_fields=TARGET_FIELDS,
                           current_field_index=current_field_index,
                           total_fields=len(TARGET_FIELDS),
                           field_mappings=field_mappings,
                           skipped_fields=skipped_fields)


def convert_field_name(field_name):
    """Alan isimlerini veritabanı alan isimlerine dönüştür"""
    field_mapping = {
        'Scenario Name': 'scenario_name',
        'Recurrence': 'recurrence',
        'Type': 'type_of_action',
        'Action': 'type_of_action',
        'Source': 'source',
        'Destination': 'destination',
        'User': 'user',
        'Email Notification': 'email_notification'
    }
    return field_mapping.get(field_name, field_name.lower().replace(' ', '_'))


@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        # Form verilerini al
        current_user.default_landing_page = request.form.get(
            'default_landing_page', 'index')
        current_user.theme = request.form.get('theme', 'light')
        current_user.sidebar_collapsed = request.form.get(
            'sidebar_collapsed') == 'on'
        current_user.show_email_notifications = request.form.get(
            'show_email_notifications') == 'on'
        current_user.items_per_page = int(
            request.form.get('items_per_page', 10))

        # Değişiklikleri kaydet
        db.session.commit()

        flash('Profil ayarlarınız başarıyla güncellendi.', 'success')
        return redirect(url_for('profile'))

    # Form için CSRF token oluştur
    form = FlaskForm()
    return render_template('profile.html', form=form)

# Login sonrası yönlendirme


@login_manager.unauthorized_handler
def unauthorized():
    return redirect(url_for('login'))


@app.route('/unauthorized')
@login_required
def unauthorized():
    return render_template('unauthorized.html'), 403


@app.route('/json-to-excel', methods=['POST'])
@login_required
def json_to_excel():
    try:
        # Get JSON data from form
        json_data = request.form.get('json_data')
        if not json_data:
            # Try to get file from request
            if 'json_file' not in request.files:
                flash('Dosya seçilmedi', 'error')
                return redirect(url_for('import_json'))
            
            file = request.files['json_file']
            if file.filename == '':
                flash('Dosya seçilmedi', 'error')
                return redirect(url_for('import_json'))
            
            if not file.filename.endswith('.json'):
                flash('Sadece JSON dosyaları desteklenir', 'error')
                return redirect(url_for('import_json'))
            
            json_data = file.read().decode('utf-8')
        
        # Parse JSON data
        data = json.loads(json_data)
        
        # Convert to DataFrame
        if isinstance(data, dict) and 'data' in data:
            df = pd.DataFrame(data['data'])
        else:
            df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Scenarios')
        
        output.seek(0)
        
        # Save scenarios to database
        for _, row in df.iterrows():
            scenario = Scenario(
                name=row.get('Scenario Name', ''),
                recurrence=row.get('Recurrence', ''),
                type=row.get('Type', ''),
                action=row.get('Action', ''),
                source=row.get('Source', ''),
                destination=row.get('Destination', ''),
                user=row.get('User', ''),
                email_notification=row.get('Email Notification', '')
            )
            db.session.add(scenario)
            
            # Save responsible person
            if 'Sorumlu Kişi' in row and row['Sorumlu Kişi']:
                responsible_person = ResponsiblePerson(
                    responsible_name=row['Sorumlu Kişi']
                )
                db.session.add(responsible_person)
        
        db.session.commit()
        
        # Return Excel file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='scenarios.xlsx'
        )

    except Exception as e:
        db.session.rollback()
        flash(f'Dönüştürme sırasında bir hata oluştu: {str(e)}', 'error')
        return redirect(url_for('import_json'))


def extract_steps(statements, attributes):
    steps = []
    email_notifications = []

    for stmt in statements:
        action = stmt.get("ActionStatement", {}).get("Action", {})
        if isinstance(action, dict):
            if "UploadAction" in action and isinstance(
                    action["UploadAction"], dict):
                upload_action = action["UploadAction"]
                steps.append({"Action Type": "Upload",
                              "Source": upload_action.get("LocalPath",
                                                          "N/A").replace('%FS.PATH%',
                                                                         attributes.get('trigger',
                                                                                        {}).get('FolderMonitorParams',
                                                                                                {}).get('Path',
                                                                                                        'N/A')),
                              "Destination": upload_action.get("RemotePath",
                                                               "N/A")})
            if "DownloadAction" in action and isinstance(
                    action["DownloadAction"], dict):
                download_action = action["DownloadAction"]
                steps.append({
                    "Action Type": "Download",
                    "Source": download_action.get("RemotePath", "N/A"),
                    "Destination": download_action.get("LocalPath", "N/A")
                })
            if "CloudUploadAction" in action and isinstance(
                    action["CloudUploadAction"], dict):
                cloud_action = action["CloudUploadAction"]
                azure_data = cloud_action.get("AzureUploadData", {})
                steps.append({
                    "Action Type": "Cloud Upload",
                    "Source": azure_data.get("LocalPath", "N/A"),
                    "Destination": azure_data.get("RemotePath", "N/A")
                })
            if "CloudDownloadAction" in action and isinstance(
                    action["CloudDownloadAction"], dict):
                cloud_action = action["CloudDownloadAction"]
                azure_data = cloud_action.get("AzureDownloadData", {})
                steps.append({
                    "Action Type": "Azure Download",
                    "Source": azure_data.get("RemotePath", cloud_action.get("RemotePath", "N/A")),
                    "Destination": azure_data.get("LocalPath", cloud_action.get("LocalPath", "N/A"))
                })

        for fail_action in stmt.get(
                "ActionStatement", {}).get(
                "IfFailedActions", []):
            if "MailAction" in fail_action:
                email_notifications.extend(
                    fail_action["MailAction"].get(
                        "AddressesTO", []))
                email_notifications.extend(
                    fail_action["MailAction"].get(
                        "AddressesCC", []))
                email_notifications.extend(
                    fail_action["MailAction"].get(
                        "AddressesBCC", []))

        if_attached = stmt.get(
            "ConditionStatement", {}).get(
            "IfAttachedStatements", [])
        sub_steps, sub_emails = extract_steps(if_attached, attributes)
        steps.extend(sub_steps)
        email_notifications.extend(sub_emails)

    return steps, email_notifications


def extract_all_scenarios(json_data):
    try:
        if isinstance(json_data, list):
            # Basit format - doğrudan liste
            return json_data
        elif isinstance(json_data, dict):
            # Karmaşık format - extract_all_scenarios kullan
            data = extract_all_scenarios(json_data)
            if data:
                # DataFrame'e dönüştür
                return data
            else:
                return None
        else:
            return None
    except Exception as e:
        print(f"Error extracting scenarios: {str(e)}")
        return None


def convert_json_to_excel(json_file_path, excel_file_path):
    try:
        # JSON dosyasını oku
        with open(json_file_path, 'r') as f:
            json_data = json.load(f)
        
        # Veri formatını kontrol et ve uygun şekilde işle
        if "data" in json_data:
            # Önce basit format için kontrol et (doğrudan senaryo verileri)
            if json_data["data"] and isinstance(json_data["data"], list) and "Scenario Name" in json_data["data"][0]:
                # Basit format - doğrudan kullan
                data = json_data["data"]
            else:
                # Karmaşık format - extract_all_scenarios kullan
                data = extract_all_scenarios(json_data)
                if data:
                    # DataFrame'e dönüştür
                    df = pd.DataFrame(data)
                    
                    # Excel dosyasını oluştur
                    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Scenarios')
                        
                        # Sütun genişliklerini ayarla
                        worksheet = writer.sheets['Scenarios']
                        for idx, col in enumerate(df.columns):
                            max_length = max(
                                df[col].astype(str).apply(len).max(),
                                len(col)
                            ) + 2
                            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 50)
                    
                    return True
                else:
                    print("Dönüştürülecek veri bulunamadı")
                    return False
        else:
            print("JSON veri yapısında 'data' alanı bulunamadı")
            return False
    except Exception as e:
        print(f"Hata oluştu: {str(e)}")
        return False

# Log Management Routes


@app.route('/log-viewer')
@login_required
def log_viewer():
    page = request.args.get('page', 1, type=int)
    per_page = current_user.items_per_page

    # Filter parameters
    scenario_id = request.args.get('scenario_id', type=int)
    log_type = request.args.get('type')
    level = request.args.get('level')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    # Base query
    query = Log.query

    # Apply filters
    if scenario_id:
        query = query.filter_by(scenario_id=scenario_id)
    if log_type:
        query = query.filter_by(log_type=log_type)
    if level:
        query = query.filter_by(level=level)
    if start_date:
        query = query.filter(
            Log.timestamp >= datetime.strptime(
                start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(
            Log.timestamp <= datetime.strptime(
                end_date, '%Y-%m-%d'))

    # Get paginated logs
    logs = query.order_by(
        Log.timestamp.desc()).paginate(
        page=page,
        per_page=per_page)

    # Get all scenarios for filter dropdown
    scenarios = Scenario.query.all()

    return render_template('log_viewer.html', logs=logs, scenarios=scenarios)


@app.route('/log-upload', methods=['GET', 'POST'])
@login_required
def log_upload():
    if request.method == 'POST':
        log_type = request.form.get('log_type')

        if log_type == 'auto':
            # Create necessary directories if they don't exist
            logs_dir = os.path.join('Uploads', 'Logs')
            archive_dir = os.path.join(logs_dir, 'archive')
            failed_dir = os.path.join(logs_dir, 'failed')

            for directory in [logs_dir, archive_dir, failed_dir]:
                if not os.path.exists(directory):
                    os.makedirs(directory)

            # Process all log files in the Uploads/Logs directory
            processed_count = 0
            failed_count = 0

            for filename in os.listdir(logs_dir):
                if filename.endswith(
                    ('.log', '.txt')) and os.path.isfile(
                    os.path.join(
                        logs_dir, filename)):
                    file_path = os.path.join(logs_dir, filename)
                    try:
                        # Read and process the log file
                        with open(file_path, 'r', encoding='utf-8') as f:
                            content = f.read()

                        # Parse log content
                        log_entries = parse_log_content(content)

                        # Create log entries in database
                        for entry in log_entries:
                            log = Log(
                                timestamp=entry.get(
                                    'timestamp',
                                    turkey_now()),
                                log_type='auto',
                                level=entry.get(
                                    'level',
                                    'INFO'),
                                message=entry.get(
                                    'message',
                                    ''),
                                source=filename,
                                parsed_data=entry,
                                created_by=current_user.id,
                                category=entry.get(
                                    'category',
                                    'general'),
                                host=entry.get(
                                    'host',
                                    ''),
                                process=entry.get(
                                    'process',
                                    ''))

                            # Try to associate with scenario if possible
                            scenario_name = extract_scenario_name(entry)
                            if scenario_name:
                                scenario = Scenario.query.filter_by(
                                    scenario_name=scenario_name).first()
                                if scenario:
                                    log.scenario_id = scenario.id

                            db.session.add(log)

                        # Move processed file to archive
                        archive_path = os.path.join(
                            archive_dir, f"{turkey_now().strftime('%Y%m%d_%H%M%S')}_{filename}")
                        os.rename(file_path, archive_path)
                        processed_count += 1

                    except Exception as e:
                        # Move failed file to failed directory
                        failed_path = os.path.join(failed_dir, filename)
                        os.rename(file_path, failed_path)
                        failed_count += 1
                        app.logger.error(
                            f"Error processing {filename}: {str(e)}")

            try:
                db.session.commit()
                flash(
                    f'İşlem tamamlandı. {processed_count} dosya işlendi, {failed_count} dosya başarısız.',
                    'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Veritabanı hatası: {str(e)}', 'error')

        else:  # Manual upload
            if 'log_file' not in request.files:
                flash('Dosya seçilmedi.', 'error')
                return redirect(request.url)

            file = request.files['log_file']
            if file.filename == '':
                flash('Dosya seçilmedi.', 'error')
                return redirect(request.url)

            if not file.filename.endswith(('.log', '.txt')):
                flash('Sadece .log ve .txt dosyaları desteklenmektedir.', 'error')
                return redirect(request.url)

            try:
                content = file.read().decode('utf-8')
                log_entries = parse_log_content(content)

                for entry in log_entries:
                    log = Log(
                        timestamp=entry.get('timestamp', turkey_now()),
                        log_type='manual',
                        level=entry.get('level', 'INFO'),
                        message=entry.get('message', ''),
                        source=file.filename,
                        parsed_data=entry,
                        created_by=current_user.id,
                        category=entry.get('category', 'general'),
                        host=entry.get('host', ''),
                        process=entry.get('process', ''))

                    # Try to associate with scenario if possible
                    scenario_name = extract_scenario_name(entry)
                    if scenario_name:
                        scenario = Scenario.query.filter_by(
                            scenario_name=scenario_name).first()
                        if scenario:
                            log.scenario_id = scenario.id

                    db.session.add(log)

                db.session.commit()
                flash('Log dosyası başarıyla yüklendi ve işlendi.', 'success')

            except Exception as e:
                db.session.rollback()
                flash(f'Hata oluştu: {str(e)}', 'error')

        return redirect(url_for('log_viewer'))

    return render_template('log_upload.html')


@app.route('/log-settings', methods=['GET', 'POST'])
@login_required
def log_settings():
    settings = LogSettings.query.first()
    if not settings:
        settings = LogSettings()
        db.session.add(settings)
        db.session.commit()

    if request.method == 'POST':
        settings.auto_import_enabled = 'auto_import_enabled' in request.form
        settings.import_directory = request.form.get('import_directory')
        settings.file_pattern = request.form.get('file_pattern')
        settings.parse_interval = int(request.form.get('parse_interval'))
        settings.retention_period = int(request.form.get('retention_period'))
        settings.max_file_size = int(request.form.get('max_file_size'))
        settings.updated_at = turkey_now()

        db.session.commit()
        flash('Log ayarları başarıyla güncellendi', 'success')
        return redirect(url_for('log_settings'))

    return render_template('log_settings.html', settings=settings)


@app.route('/export-logs', methods=['GET'])
@login_required
def export_logs():
    # Get filter parameters
    scenario_id = request.args.get('scenario_id', type=int)
    log_type = request.args.get('type')
    level = request.args.get('level')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    # Base query
    query = Log.query

    # Apply filters
    if scenario_id:
        query = query.filter_by(scenario_id=scenario_id)
    if log_type:
        query = query.filter_by(log_type=log_type)
    if level:
        query = query.filter_by(level=level)
    if start_date:
        query = query.filter(
            Log.timestamp >= datetime.strptime(
                start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(
            Log.timestamp <= datetime.strptime(
                end_date, '%Y-%m-%d'))

    # Get all logs based on filters
    logs = query.order_by(Log.timestamp.desc()).all()

    # Create a temporary Excel file
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_file.close()

    # Create a Pandas DataFrame from logs
    log_data = []
    for log in logs:
        scenario_name = log.scenario.scenario_name if log.scenario else ''
        created_by = log.user.username if log.user else ''

        log_entry = {
            'Tarih': log.timestamp,
            'Seviye': log.level,
            'Tür': log.log_type,
            'Kategori': log.category,
            'Mesaj': log.message,
            'Kaynak': log.source,
            'Host': log.host,
            'Süreç': log.process,
            'Senaryo': scenario_name,
            'Oluşturan': created_by
        }
        log_data.append(log_entry)

    if not log_data:
        flash('Dışa aktarılacak log kaydı bulunamadı.', 'warning')
        return redirect(url_for('log_viewer'))

    df = pd.DataFrame(log_data)

    # Create Excel workbook
    with pd.ExcelWriter(temp_file.name, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Logs', index=False)

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Logs']

        # Add header formatting
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#007bff',
            'color': 'white',
            'border': 1
        })

        # Apply formatting
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)

        # Auto-adjust columns width
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(col)
            )
            worksheet.set_column(idx, idx, max_length + 2)

    # Send file
    return send_file(
        temp_file.name,
        as_attachment=True,
        download_name=f'logs_export_{turkey_now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def parse_log_content(content):
    """
    Parse log content and extract relevant information.
    Returns a list of dictionaries containing parsed log entries.
    """
    parsed_entries = []
    # Split content into lines
    lines = content.split('\n')
    for line in lines:
        if not line.strip():
            continue
        try:
            # Try to parse as JSON first
            entry = json.loads(line)
        except json.JSONDecodeError:
            # If not JSON, try to parse using regular expressions
            entry = parse_log_line(line)
        if entry:
            # Parse timestamp if it's a string
            if 'timestamp' in entry and isinstance(entry['timestamp'], str):
                try:
                    entry['timestamp'] = date_parser.parse(entry['timestamp'])
                except Exception:
                    entry['timestamp'] = None
            # Try to match with scenario
            scenario_name = extract_scenario_name(entry)
            if scenario_name:
                scenario = Scenario.query.filter(
                    Scenario.scenario_name.like(f"%{scenario_name}%")).first()
                if scenario:
                    entry['scenario_id'] = scenario.id
            parsed_entries.append(entry)
    return parsed_entries


def parse_log_line(line):
    """
    Parse a single log line using regular expressions.
    Returns a dictionary containing extracted information.
    """
    # Common log patterns
    patterns = [
        # Timestamp Level Message
        r'^(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}(?:\.\d+)?)\s+(\w+)\s+(.+)$',
        # [Timestamp] Level: Message
        r'^\([^\]]+\)\s+(\w+):\s+(.+)$',
        # Message only
        r'^(.+)$'
    ]

    for pattern in patterns:
        match = re.match(pattern, line)
        if match:
            groups = match.groups()
            if len(groups) == 3:
                # Parse timestamp string to datetime
                try:
                    timestamp = date_parser.parse(groups[0])
                except Exception:
                    timestamp = None
                return {
                    'timestamp': timestamp,
                    'level': groups[1],
                    'message': groups[2]
                }
            elif len(groups) == 1:
                return {
                    'message': groups[0],
                    'level': 'INFO'
                }
    return None


def extract_scenario_name(entry):
    """
    Extract scenario name from log entry.
    Returns the scenario name if found, None otherwise.
    """
    message = entry.get('message', '').lower()

    # Common patterns to identify scenario names
    patterns = [
        r'scenario[:\s]+([^\s]+)',
        r'senaryo[:\s]+([^\s]+)',
        r'executing[:\s]+([^\s]+)',
        r'running[:\s]+([^\s]+)'
    ]

    for pattern in patterns:
        match = re.search(pattern, message)
        if match:
            return match.group(1)

    return None


@app.route('/api/logs/<int:log_id>')
@login_required
def get_log_details(log_id):
    log = Log.query.get_or_404(log_id)
    return jsonify({
        'id': log.id,
        'timestamp': log.timestamp.isoformat(),
        'level': log.level,
        'message': log.message,
        'source': log.source,
        'parsed_data': log.parsed_data or {},
        'category': log.category,
        'host': log.host,
        'process': log.process,
        'log_type': log.log_type
    })


@app.route('/import-scenarios/preview', methods=['POST'])
@csrf.exempt
def import_scenarios_preview():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'Lütfen bir dosya seçin'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'Dosya seçilmedi'})

    if not file.filename.endswith(('.xlsx', '.csv')):
        return jsonify(
            {'success': False, 'message': 'Sadece Excel (.xlsx) veya CSV dosyaları desteklenir'})

    temp_file = None
    try:
        # Geçici dizin oluştur
        temp_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'temp')
        os.makedirs(temp_dir, exist_ok=True)

        # Güvenli dosya adı oluştur
        safe_filename = secure_filename(file.filename)
        temp_file = os.path.join(temp_dir, safe_filename)

        # Dosyayı kaydet
        file.save(temp_file)

        # Dosyanın varlığını kontrol et
        if not os.path.exists(temp_file):
            return jsonify(
                {'success': False, 'message': 'Dosya kaydedilemedi'})

        # Geçici dosya yolunu session'da sakla
        session['temp_file_path'] = temp_file

        # Dosyayı oku
        try:
            if file.filename.endswith('.csv'):
                df = pd.read_csv(temp_file, encoding='utf-8')
            else:
                df = pd.read_excel(temp_file)
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Dosya okuma hatası: {str(e)}'
            })

        # Gerekli alanları kontrol et
        required_fields = [
            'Scenario Name',
            'Recurrence',
            'Action Type',
            'Source',
            'Destination']
        missing_fields = [
            field for field in required_fields if field not in df.columns]

        if missing_fields:
            return jsonify({
                'success': False,
                'message': f'Eksik zorunlu alanlar: {", ".join(missing_fields)}'
            })

        # En yüksek Custom ID'yi bul ve 100'ü çıkar
        max_custom_id = 0
        for custom_id in df['Custom ID']:
            try:
                custom_id_num = int(custom_id)
                max_custom_id = max(max_custom_id, custom_id_num)
            except (ValueError, TypeError):
                continue

        # Toplam senaryo sayısını hesapla
        total_scenarios = len(df['Custom ID'].unique())

        # Custom ID'ye göre grupla ve her grup için geçerlilik kontrolü yap
        active_count = 0
        passive_count = 0

        for custom_id in df['Custom ID'].unique():
            if pd.isna(custom_id):
                continue

            group = df[df['Custom ID'] == custom_id]

            # Gruptaki her satır için zorunlu alanları kontrol et
            is_valid = True
            for _, row in group.iterrows():
                if not all(str(row[field]).strip()
                           for field in required_fields):
                    is_valid = False
                    break

            # Status kontrolü
            status = str(group.iloc[0].get('Status', '')).lower().strip()
            if status == 'aktif':
                status = 'active'
            elif status == 'pasif':
                status = 'passive'
            else:
                status = 'active'  # Varsayılan olarak aktif

            if is_valid:
                if status == 'active':
                    active_count += 1
                else:
                    passive_count += 1

        # Önizleme için ilk 10 senaryoyu hazırla
        preview_data = []
        for _, row in df.head(10).iterrows():
            # Status dönüşümü
            status = str(row.get('Status', '')).lower().strip()
            if status == 'aktif':
                status = 'active'
            elif status == 'pasif':
                status = 'passive'
            else:
                status = 'active'  # Varsayılan olarak aktif

            preview_data.append({
                'custom_id': str(row.get('Custom ID', '')),
                'scenario_name': str(row.get('Scenario Name', '')),
                'recurrence': str(row.get('Recurrence', '')),
                'repeat_rate': str(row.get('Repeat Rate', '')),
                'type_of_action': str(row.get('Action Type', '')),
                'type': str(row.get('Type', '')),
                'step_count': str(row.get('Step Count', '')),
                'source': str(row.get('Source', '')),
                'destination': str(row.get('Destination', '')),
                'status': status,
                'folder_id': str(row.get('Folder ID', '')),
                'department': str(row.get('Department', ''))
            })

        return jsonify({
            'success': True,
            'total': total_scenarios,
            'active': active_count,
            'passive': passive_count,
            'preview': preview_data
        })

    except Exception as e:
        # Hata durumunda geçici dosyayı temizle
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except BaseException:
                pass
        session.pop('temp_file_path', None)
        return jsonify({
            'success': False,
            'message': f'Dosya işlenirken hata oluştu: {str(e)}'
        })


@app.route('/import-scenarios/import', methods=['POST'])
@login_required
@csrf.exempt
def import_scenarios_data():
    try:
        # Get the temporary file path from session
        temp_file_path = session.get('temp_file_path')
        if not temp_file_path or not os.path.exists(temp_file_path):
            return jsonify({
                'success': False,
                'message': 'Geçici dosya bulunamadı'
            }), 400

        # Read the file based on its extension
        file_extension = os.path.splitext(temp_file_path)[1].lower()
        try:
            if file_extension == '.xlsx':
                df = pd.read_excel(temp_file_path, engine='openpyxl')
            elif file_extension == '.xls':
                df = pd.read_excel(temp_file_path, engine='xlrd')
            elif file_extension == '.csv':
                # Try different encodings for CSV
                try:
                    df = pd.read_csv(temp_file_path, encoding='utf-8')
                except UnicodeDecodeError:
                    try:
                        df = pd.read_csv(temp_file_path, encoding='latin1')
                    except:
                        df = pd.read_csv(temp_file_path, encoding='cp1252')
            else:
                return jsonify({
                    'success': False,
                    'message': f'Desteklenmeyen dosya formatı: {file_extension}'
                }), 400
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Dosya okunamadı: {str(e)}'
            }), 400

        # Sütun isimlerini düzelt
        column_mapping = {
            'Scenario Name': 'scenario_name',
            'Action Type': 'type_of_action',
            'Recurrence': 'recurrence',
            'Repeat Rate': 'repeat_rate',
            'Source': 'source',
            'Destination': 'destination',
            'Email Notification': 'email_notification',
            'Status': 'status',
            'Folder ID': 'folder_id',
            'Type': 'type',
            'Step Count': 'step_count',
            'Department': 'department'
        }

        # Sütun isimlerini dönüştür
        df = df.rename(columns={k: v for k, v in column_mapping.items() if k in df.columns})

        # Process each row
        for _, row in df.iterrows():
            # Check if scenario exists in responsible_persons_new
            responsible_person = ResponsiblePerson.query.filter_by(
                scenario_name=row['scenario_name']
            ).first()
            
            # If no match found, set default responsible person as Admin
            responsible_person_name = responsible_person.responsible_name if responsible_person else 'Admin'
            
            # Convert step_count to integer if it exists and is not empty
            step_count = None
            if 'step_count' in row and row['step_count'] != '':
                try:
                    step_count = int(float(row['step_count']))
                except (ValueError, TypeError):
                    step_count = 0
            
            # Status dönüşümü
            status = str(row.get('status', '')).lower().strip()
            if status == 'aktif':
                status = 'active'
            elif status == 'pasif':
                status = 'passive'
            else:
                status = 'active'  # Varsayılan olarak aktif
            
            # Create new scenario
            scenario = Scenario(
                scenario_name=str(row['scenario_name']),
                type_of_action=str(row['type_of_action']),
                user=current_user.username,  # Use current user as default
                recurrence=str(row['recurrence']),
                repeat_rate=str(row.get('repeat_rate', '')),
                source=str(row['source']),
                destination=str(row['destination']),
                email_notification=str(row.get('email_notification', '')),
                responsible_person=responsible_person_name,  # Use the responsible person from responsible_persons_new or default to Admin
                created_by=current_user.id,
                status=status,  # Use the converted status
                group_id=str(uuid.uuid4()),
                folder_id=str(row.get('folder_id', '')),
                type=str(row.get('type', '')),
                step_count=step_count,
                department=str(row.get('department', ''))
            )
            db.session.add(scenario)
        db.session.commit()

        # Clean up temporary file
        try:
            os.remove(temp_file_path)
            session.pop('temp_file_path', None)
        except Exception as e:
            print(f"Error removing temporary file: {str(e)}")

        return jsonify({
            'success': True,
            'message': 'Senaryolar başarıyla içe aktarıldı'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error importing scenarios: {str(e)}'
        }), 500


@app.route('/import-scenarios/get-converted-file/<filename>')
@login_required
@csrf.exempt
def get_converted_file(filename):
    """Dönüştürülmüş Excel dosyasını indirme"""
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        flash('Dosya bulunamadı', 'error')
        return redirect(url_for('import_scenarios'))

@app.route('/admin/assign-role', methods=['POST'])
@login_required
@csrf.exempt
def assign_role():
    if not current_user.has_permission('role_management'):
        return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok.'})
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        role_id = data.get('role_id')
        if not user_id or not role_id:
            return jsonify({'success': False, 'message': 'Gerekli bilgiler eksik'})
        user = User.query.get(user_id)
        role = Role.query.get(role_id)
        if not user or not role:
            return jsonify({'success': False, 'message': 'Kullanıcı veya rol bulunamadı'})
        if role not in user.roles:
            user.roles.append(role)
            # Log role assignment with UTC timestamp
            auth_log = AuthLog(
                username=current_user.username,
                action='role_assigned',
                status='success',
                ip_address=request.remote_addr,
                details=f'Role {role.name} assigned to user {user.username}',
                target_user=user.username,
                target_role=role.name,
                performed_by=current_user.username,
                timestamp=turkey_now()  # Explicitly use UTC
            )
            db.session.add(auth_log)
            db.session.commit()
            current_app.logger.info(f"Role {role.name} assigned to user {user.username} by {current_user.username}")
            return jsonify({'success': True, 'message': 'Rol başarıyla atandı'})
        else:
            return jsonify({'success': False, 'message': 'Kullanıcı zaten bu role sahip'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error assigning role: {str(e)}")
        return jsonify({'success': False, 'message': f'Bir hata oluştu: {str(e)}'})

@app.route('/admin/revoke-role', methods=['POST'])
@login_required
@csrf.exempt
def revoke_role():
    if not current_user.has_permission('role_management'):
        return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok.'})
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        role_id = data.get('role_id')
        if not user_id or not role_id:
            return jsonify({'success': False, 'message': 'Gerekli bilgiler eksik'})
        user = User.query.get(user_id)
        role = Role.query.get(role_id)
        if not user or not role:
            return jsonify({'success': False, 'message': 'Kullanıcı veya rol bulunamadı'})
        if role in user.roles:
            user.roles.remove(role)
            # Log role revocation with UTC timestamp
            auth_log = AuthLog(
                username=current_user.username,
                action='role_revoked',
                status='success',
                ip_address=request.remote_addr,
                details=f'Role {role.name} revoked from user {user.username}',
                target_user=user.username,
                target_role=role.name,
                performed_by=current_user.username,
                timestamp=turkey_now()  # Explicitly use UTC
            )
            db.session.add(auth_log)
            db.session.commit()
            current_app.logger.info(f"Role {role.name} revoked from user {user.username} by {current_user.username}")
            return jsonify({'success': True, 'message': 'Rol başarıyla kaldırıldı'})
        else:
            return jsonify({'success': False, 'message': 'Kullanıcı bu role sahip değil'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error revoking role: {str(e)}")
        return jsonify({'success': False, 'message': f'Bir hata oluştu: {str(e)}'})

@app.route('/update-scenario-responsible/<int:scenario_id>', methods=['POST'])
@login_required
def update_scenario_responsible(scenario_id):
    try:
        # Get the scenario
        scenario = Scenario.query.get_or_404(scenario_id)
        
        # Get the new responsible person from either form data or JSON
        new_responsible = None
        if request.is_json:
            data = request.get_json()
            new_responsible = data.get('responsible_person')
        else:
            new_responsible = request.form.get('responsible_person')
            
        if not new_responsible:
            return jsonify({
                'success': False,
                'message': 'Sorumlu kişi belirtilmedi'
            }), 400

        # Store old responsible for logging
        old_responsible = scenario.responsible_person

        # Update the scenario's responsible person
        scenario.responsible_person = new_responsible
        
        # Update or create responsible person record
        responsible_person = ResponsiblePerson.query.filter_by(
            scenario_name=scenario.scenario_name
        ).first()
        
        if responsible_person:
            responsible_person.responsible_name = new_responsible
            responsible_person.updated_at = turkey_now()
        else:
            responsible_person = ResponsiblePerson(
                scenario_name=scenario.scenario_name,
                responsible_name=new_responsible
            )
            db.session.add(responsible_person)

        # Log the change
        log = AuthLog(
            username=current_user.username,
            action='responsible_updated',
            status='success',
            ip_address=request.remote_addr,
            details=f'Senaryo sorumlusu güncellendi: {old_responsible} -> {new_responsible}',
            target_scenario=scenario.scenario_name,
            performed_by=current_user.username
        )
        db.session.add(log)

        # Commit the changes
        db.session.commit()
            
        return jsonify({
            'success': True,
            'message': 'Sorumlu kişi başarıyla güncellendi'
        })
            
    except Exception as e:
        db.session.rollback()
        print(f"Error updating scenario responsible: {str(e)}")  # Log the error
        return jsonify({
            'success': False,
            'message': f'Güncelleme başarısız oldu: {str(e)}'
        }), 500
            
@app.route('/get-responsible-persons')
@login_required
def get_responsible_persons():
    try:
        # Get all users with full_name
        users = User.query.all()
        
        # Create dictionary with full_name or username
        responsible_persons = {}
        for user in users:
            display_name = user.full_name if user.full_name else user.username
            responsible_persons[display_name] = display_name
        
        return jsonify({
            'success': True,
            'responsible_persons': responsible_persons
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        })

@app.route('/scenarios')
@login_required
def scenarios():
    # Get all scenarios
    scenarios = Scenario.query.all()
    
    # Check if user has permission to edit responsible person or is a Scenario Manager
    can_edit_responsible = (
        current_user.has_permission('edit_responsible_person') or
        any(role.name == 'Senaryo Yöneticisi' for role in current_user.roles)
    )
    
    return render_template('scenarios.html', 
                         scenarios=scenarios,
                         can_edit_responsible=can_edit_responsible)


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        email = request.form['email']
        
        if User.query.filter_by(username=username).first():
            flash('Bu kullanıcı adı zaten kullanılıyor.', 'danger')
            return redirect(url_for('register'))
        
        if User.query.filter_by(email=email).first():
            flash('Bu e-posta adresi zaten kullanılıyor.', 'danger')
            return redirect(url_for('register'))
        
        user = User(username=username, email=email)
        user.set_password(password)
        db.session.add(user)
        
        # Log the user creation
        log = AuthLog(
            username=username,
            action='user_created',
            status='success',
            ip_address=request.remote_addr,
            details=f'Yeni kullanıcı oluşturuldu: {username}',
            performed_by=current_user.username if not current_user.is_anonymous else 'system'
        )
        db.session.add(log)

        try:
            db.session.commit()
            flash('Kayıt başarılı! Şimdi giriş yapabilirsiniz.', 'success')
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            flash(f'Kayıt sırasında bir hata oluştu: {str(e)}', 'danger')
            return redirect(url_for('register'))
    
    return render_template('register.html')


def test_ad_connection(settings):
    """Test Active Directory connection with given settings"""
    results = {
        'success': False,
        'steps': [],
        'error': None
    }
    
    try:
        # Validate required settings first
        if not settings.ad_server:
            raise ValueError("Server address is required")
        if not settings.ad_username:
            raise ValueError("Username is required")
        if not settings.ad_password:
            raise ValueError("Password is required")
        if not settings.ad_domain:
            raise ValueError("Domain is required")
        if not settings.ad_base_dn:
            raise ValueError("Base DN is required")
        
        # Step 1: Server connection test
        results['steps'].append({
            'step': 'Server Connection',
            'status': 'pending',
            'details': f'Attempting to connect to {settings.ad_server}:{settings.ad_port}'
        })
        
        server = Server(
            settings.ad_server,
            port=settings.ad_port,
            use_ssl=settings.ad_use_ssl,
            get_info=ALL
        )
        
        # Format username for NTLM authentication
        formatted_username = f"{settings.ad_domain}\\{settings.ad_username}"
        
        # Step 2: Authentication test
        results['steps'].append({
            'step': 'Authentication',
            'status': 'pending',
            'details': f'Attempting to authenticate with username: {formatted_username}'
        })
        
        # Create connection without auto_bind
        conn = Connection(
            server,
            user=formatted_username,
            password=settings.ad_password,
            authentication=NTLM
        )
        
        # Try to bind explicitly
        if not conn.bind():
            raise ValueError(f"Bind failed: {conn.result}")
        
        # Step 3: Base DN test
        results['steps'].append({
            'step': 'Base DN Test',
            'status': 'pending',
            'details': f'Testing base DN: {settings.ad_base_dn}'
        })
        
        search_result = conn.search(
            search_base=settings.ad_base_dn,
            search_filter='(objectClass=*)',
            search_scope=SUBTREE,
            attributes=['*'],
            size_limit=1
        )
        
        if not search_result:
            raise ValueError(f"Base DN search failed: {conn.result}")
        
        # Step 4: Admin group test
        results['steps'].append({
            'step': 'Admin Group Test',
            'status': 'pending',
            'details': f'Testing admin group: {settings.ad_admin_group}'
        })
        
        if not settings.ad_admin_group:
            raise ValueError("Admin group is required")
        
        admin_group_filter = settings.ad_group_filter.replace('{group}', settings.ad_admin_group)
        search_result = conn.search(
            search_base=settings.ad_base_dn,
            search_filter=admin_group_filter,
            search_scope=SUBTREE,
            attributes=['*'],
            size_limit=1
        )
        
        if not search_result:
            raise ValueError(f"Admin group search failed: {conn.result}")
        
        # Update all steps to success
        for step in results['steps']:
            step['status'] = 'success'
            step['details'] += ' - Success'
        
        results['success'] = True
        conn.unbind()
        
    except Exception as e:
        # Update the last pending step with error
        for step in reversed(results['steps']):
            if step['status'] == 'pending':
                step['status'] = 'error'
                step['details'] += f' - Failed: {str(e)}'
                break
        
        results['error'] = str(e)
    
    return results

@app.route('/test-ad-connection', methods=['POST'])
@login_required
def test_ad_connection_route():
    if not current_user.is_admin:
        return jsonify({
            'success': False,
            'message': 'You do not have permission to test AD settings'
        }), 403
    
    try:
        settings = SystemSettings.query.first()
        if not settings:
            return jsonify({
                'success': False,
                'message': 'System settings not found'
            }), 404
        
        results = test_ad_connection(settings)
        return jsonify(results)
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error testing AD connection: {str(e)}'
        }), 500

def create_default_permissions():
    default_permissions = [
        ('manage_users', 'Can manage users'),
        ('manage_roles', 'Can manage roles'),
        ('manage_scenarios', 'Can manage scenarios'),
        ('view_logs', 'Can view logs'),
        ('manage_logs', 'Can manage logs'),
        ('view_auth_logs', 'Can view authentication logs'),
        ('manage_system', 'Can manage system settings'),
        ('export_data', 'Can export data'),
        ('import_data', 'Can import data'),
        ('manage_change_requests', 'Can manage change requests'),
        ('view_change_requests', 'Can view change requests'),
        ('manage_responsible_persons', 'Can manage responsible persons')
    ]
    
    for name, description in default_permissions:
        permission = Permission.query.filter_by(name=name).first()
        if not permission:
            permission = Permission(name=name, description=description)
            db.session.add(permission)
    
    db.session.commit()

def create_default_roles():
    # Create admin role if it doesn't exist
    admin_role = Role.query.filter_by(name='admin').first()
    if not admin_role:
        admin_role = Role(
            name='admin',
            description='Administrator role with all permissions',
            is_admin=True
        )
        db.session.add(admin_role)
        db.session.commit()

    # Create user manager role if it doesn't exist
    user_manager_role = Role.query.filter_by(name='user_manager').first()
    if not user_manager_role:
        user_manager_role = Role(
            name='user_manager',
            description='Can manage users and their roles'
        )
        db.session.add(user_manager_role)
    db.session.commit()

    # Create scenario manager role if it doesn't exist
    scenario_manager_role = Role.query.filter_by(name='scenario_manager').first()
    if not scenario_manager_role:
        scenario_manager_role = Role(
            name='scenario_manager',
            description='Can manage scenarios'
        )
        db.session.add(scenario_manager_role)
        db.session.commit()

    # Create log manager role if it doesn't exist
    log_manager_role = Role.query.filter_by(name='log_manager').first()
    if not log_manager_role:
        log_manager_role = Role(
            name='log_manager',
            description='Can manage logs and view auth logs'
        )
        db.session.add(log_manager_role)
        db.session.commit()

    # Create responsible update role if it doesn't exist
    responsible_update_role = Role.query.filter_by(name='responsible_update').first()
    if not responsible_update_role:
        responsible_update_role = Role(
            name='responsible_update',
            description='Can update scenario responsible persons'
        )
        db.session.add(responsible_update_role)
        db.session.commit()

    # Assign permissions to roles
    permissions = {
        'admin': [
            'manage_users', 'manage_roles', 'manage_scenarios',
            'view_logs', 'manage_logs', 'view_auth_logs',
            'manage_system', 'export_data', 'import_data',
            'manage_change_requests', 'view_change_requests',
            'manage_responsible_persons'
        ],
        'user_manager': ['manage_users', 'manage_roles'],
        'scenario_manager': ['manage_scenarios', 'export_data', 'import_data'],
        'log_manager': ['view_logs', 'manage_logs', 'view_auth_logs'],
        'responsible_update': ['manage_responsible_persons']
    }

    for role_name, permission_names in permissions.items():
        role = Role.query.filter_by(name=role_name).first()
        if role:
            # Clear existing permissions
            role.permissions = []
            # Add new permissions
            for permission_name in permission_names:
                permission = Permission.query.filter_by(name=permission_name).first()
                if permission:
                    role.permissions.append(permission)
        db.session.commit()

def initialize_permissions():
    """Initialize default permissions and roles"""
    create_default_permissions()
    create_default_roles()

@app.route('/admin/roles', methods=['GET', 'POST'])
@login_required
def manage_roles():
    # Rol yönetimi için yetki kontrolü
    if not current_user.has_permission('role_management'):
        flash('Bu sayfaya erişim izniniz yok.', 'error')
        return redirect(url_for('unauthorized'))

    if request.method == 'POST':
        try:
            role_id = request.form.get('role_id')
            permission_names = request.form.getlist('permissions')
            role = Role.query.get(role_id)
            if role:
                # Clear existing permissions
                role.permissions = []
                # Add new permissions
                for permission_name in permission_names:
                    permission = Permission.query.filter_by(name=permission_name).first()
                    if permission:
                        role.permissions.append(permission)
                db.session.commit()
                flash('Rol izinleri başarıyla güncellendi.', 'success')
            else:
                flash('Rol bulunamadı.', 'error')
        except Exception as e:
            db.session.rollback()
            flash(f'Rol izinleri güncellenirken bir hata oluştu: {str(e)}', 'error')

    roles = Role.query.all()
    permissions = Permission.query.all()
    users = User.query.all()
    return render_template('role_management.html', roles=roles, permissions=permissions, users=users)

@app.route('/import-scenarios/clear-all', methods=['POST'])
@login_required
@csrf.exempt
def clear_all_scenarios():
    try:
        # First delete records from related tables
        # Delete from change_request table
        db.session.execute('DELETE FROM change_request')
        # Delete from log table where scenario_id is not null
        db.session.execute('DELETE FROM log WHERE scenario_id IS NOT NULL')
        # Delete from auth_logs table where target_scenario is not null
        db.session.execute('DELETE FROM auth_logs WHERE target_scenario IS NOT NULL')
        # Finally delete from scenario table
        db.session.execute('DELETE FROM scenario')
        # Note: We don't delete from responsible_persons_new table
        # to keep the responsible person information for future scenario imports
        db.session.commit()
        return jsonify({
            'success': True,
            'message': 'Tüm senaryolar başarıyla silindi'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error clearing scenarios: {str(e)}'
        }), 500

@app.route('/logs')
@login_required
def auth_logs():
    # Check if user has permission to view auth logs
    if not current_user.is_admin and not current_user.has_permission('view_auth_logs'):
        flash('Bu sayfaya erişim izniniz yok.', 'error')
        return redirect(url_for('unauthorized'))

    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    username = request.args.get('username')
    action = request.args.get('action')
    status = request.args.get('status')
    target_user = request.args.get('target_user')
    target_role = request.args.get('target_role')
    target_scenario = request.args.get('target_scenario')
    performed_by = request.args.get('performed_by')

    # Build query
    query = AuthLog.query

    if start_date:
        query = query.filter(AuthLog.timestamp >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(AuthLog.timestamp <= datetime.strptime(end_date, '%Y-%m-%d'))
    if username:
        query = query.filter(AuthLog.username.ilike(f'%{username}%'))
    if action:
        query = query.filter(AuthLog.action == action)
    if status:
        query = query.filter(AuthLog.status == status)
    if target_user:
        query = query.filter(AuthLog.target_user.ilike(f'%{target_user}%'))
    if target_role:
        query = query.filter(AuthLog.target_role.ilike(f'%{target_role}%'))
    if target_scenario:
        query = query.filter(AuthLog.target_scenario.ilike(f'%{target_scenario}%'))
    if performed_by:
        query = query.filter(AuthLog.performed_by.ilike(f'%{performed_by}%'))

    # Get unique values for filters
    usernames = db.session.query(AuthLog.username.distinct()).all()
    actions = db.session.query(AuthLog.action.distinct()).all()
    statuses = db.session.query(AuthLog.status.distinct()).all()
    target_users = db.session.query(AuthLog.target_user.distinct()).all()
    target_roles = db.session.query(AuthLog.target_role.distinct()).all()
    target_scenarios = db.session.query(AuthLog.target_scenario.distinct()).all()
    performed_bys = db.session.query(AuthLog.performed_by.distinct()).all()

    # Get pagination parameters
    page = request.args.get('page', 1, type=int)
    per_page = 50  # Show 50 logs per page

    # Order by timestamp descending (newest first)
    query = query.order_by(AuthLog.timestamp.desc())

    # Get paginated results
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    logs = pagination.items

    return render_template(
        'auth_logs.html',
        logs=logs,
        pagination=pagination,
        start_date=start_date,
        end_date=end_date,
        username=username,
        action=action,
        status=status,
        target_user=target_user,
        target_role=target_role,
        target_scenario=target_scenario,
        performed_by=performed_by,
        usernames=[u[0] for u in usernames if u[0]],
        actions=[a[0] for a in actions if a[0]],
        statuses=[s[0] for s in statuses if s[0]],
        target_users=[u[0] for u in target_users if u[0]],
        target_roles=[r[0] for r in target_roles if r[0]],
        target_scenarios=[s[0] for s in target_scenarios if s[0]],
        performed_bys=[p[0] for p in performed_bys if p[0]]
    )

# --- DÜZELTME BAŞLANGICI ---
@app.route('/update-admin-password', methods=['POST'])
@login_required
def update_admin_password():
    if not current_user.is_admin:
        flash('Bu işlem için yetkiniz yok.', 'error')
        return redirect(url_for('index'))

    try:
        # Get admin user
        admin = User.query.filter_by(is_admin=True).first()
        if not admin:
            flash('Admin kullanıcısı bulunamadı.', 'error')
            return redirect(url_for('index'))

        # Update password
        new_password = request.form.get('new_password')
        if not new_password:
            flash('Yeni şifre boş olamaz.', 'error')
            return redirect(url_for('index'))

        admin.set_password(new_password)
        db.session.commit()

        # Log the password change
        log_auth_event(
            username=current_user.username,
            action='password_change',
            status='success',
            ip_address=request.remote_addr,
            user_agent=request.user_agent.string,
            details='Admin password updated',
            target_user=admin.username
        )

        flash('Admin şifresi başarıyla güncellendi.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Şifre güncellenirken bir hata oluştu: {str(e)}', 'error')

    return redirect(url_for('index'))

# Add a simple form to update admin password
@app.route('/admin-password')
@login_required
def admin_password_form():
    if not current_user.is_admin:
        flash('Bu sayfaya erişim izniniz yok.', 'error')
        return redirect(url_for('index'))
    
    return render_template('admin_password.html')
# --- DÜZELTME SONU ---

@app.route('/admin/get-user-roles/<int:user_id>', methods=['GET'])
@login_required
def get_user_roles(user_id):
    # Check if user has admin role or role management permission
    if not current_user.is_admin and not current_user.has_permission('manage_roles'):
        return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok.'})

    user = db.session.get(User, user_id)
    if not user:
        return jsonify({'success': False, 'message': 'Kullanıcı bulunamadı'})

    # Get user's roles
    roles = [{'id': role.id, 'name': role.name} for role in user.roles]

    return jsonify({'success': True, 'roles': roles})

@app.route('/rest_api_connections')
@login_required
def rest_api_connections():
    # Check if user has admin or scenario_manager role
    if not current_user.is_admin and 'scenario_manager' not in [role.name for role in current_user.roles]:
        flash('Bu sayfaya erişim yetkiniz bulunmamaktadır.', 'error')
        return redirect(url_for('index'))

    # Tüm bağlantıları getir
    connections = RestApiConnection.query.all()
    return render_template('rest_api_connections.html', connections=connections)

@app.route('/rest-api/add-connection', methods=['POST'])
@login_required
@csrf.exempt
def add_rest_api_connection():
    try:
        # Form verilerini al
        name = request.form.get('name')
        ip_address = request.form.get('ip_address')
        port = request.form.get('port')
        username = request.form.get('username')
        password = request.form.get('password')
        
        # Gerekli alanları kontrol et
        if not all([name, ip_address, port, username, password]):
            return jsonify({
                'success': False,
                'error': 'Tüm alanları doldurunuz'
            }), 400
        
        # Yeni bağlantı oluştur
        connection = RestApiConnection(
            name=name,
            ip_address=ip_address,
            port=port,
            username=username,
            password=password
        )
        db.session.add(connection)
        db.session.commit()
        
        # Log ekle
        log = Log(
            log_type='system',
            level='INFO',
            message=f'REST API bağlantısı eklendi: {name}',
            source='rest_api',
            created_by=current_user.id
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Bağlantı başarıyla eklendi'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/rest-api/edit-connection/<int:connection_id>', methods=['POST'])
@login_required
@csrf.exempt
def edit_rest_api_connection(connection_id):
    try:
        connection = RestApiConnection.query.get_or_404(connection_id)
        
        # Form verilerini al
        name = request.form.get('name')
        ip_address = request.form.get('ip_address')
        port = request.form.get('port')
        username = request.form.get('username')
        password = request.form.get('password')
        
        # Gerekli alanları kontrol et
        if not all([name, ip_address, port, username]):
            return jsonify({
                'success': False,
                'error': 'Tüm alanları doldurunuz'
            }), 400
        
        # Bağlantıyı güncelle
        connection.name = name
        connection.ip_address = ip_address
        connection.port = port
        connection.username = username
        
        # Şifre sadece değiştirilmek istenirse güncelle
        if password:
            connection.password = password
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Bağlantı başarıyla güncellendi'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Bağlantı güncellenirken bir hata oluştu: {str(e)}'
        }), 500

@app.route('/rest-api/connections', methods=['GET'])
@login_required
def get_rest_api_connections():
    try:
        connections = RestApiConnection.query.all()
        return jsonify({
            'success': True,
            'connections': [{
                'id': conn.id,
                'name': conn.name,
                'ip_address': conn.ip_address,
                'port': conn.port,
                'username': conn.username,
                'is_active': conn.is_active,
                'auth_token': conn.auth_token,
                'last_connection': conn.last_connection.strftime('%Y-%m-%d %H:%M:%S') if conn.last_connection else None
            } for conn in connections]
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Bağlantılar alınırken bir hata oluştu: {str(e)}'
        }), 500

@app.route('/rest-api/delete-connection/<int:connection_id>', methods=['POST'])
@login_required
@csrf.exempt
def delete_rest_api_connection(connection_id):
    try:
        connection = RestApiConnection.query.get_or_404(connection_id)
        db.session.delete(connection)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/rest-api/test-connection/<int:connection_id>', methods=['POST'])
@login_required
@csrf.exempt
def test_rest_api_connection(connection_id):
    try:
        connection = RestApiConnection.query.get_or_404(connection_id)
        
        # API endpoint'ini oluştur
        api_url = f"http://{connection.ip_address}:{connection.port}/admin/v1/authentication"
        
        # İstek verilerini hazırla
        payload = {
            "userName": connection.username,
            "password": connection.password,
            "authType": "EFT"
        }
        
        headers = {
            'Content-Type': 'application/json'
        }
        
        print(f"API isteği gönderiliyor: {api_url}")  # Debug log
        
        # API'ye istek at
        response = requests.post(api_url, json=payload, headers=headers, verify=False)
        
        print(f"API yanıtı: {response.status_code} - {response.text}")  # Debug log
        
        # Yanıtı kontrol et
        if response.status_code == 200:
            try:
                data = response.json()
                print(f"API yanıt verisi: {data}")  # Debug log
                
                if 'authToken' in data:
                    # Bağlantı başarılı, son bağlantı zamanını ve token'ı güncelle
                    connection.last_connection = turkey_now()
                    connection.auth_token = data['authToken']
                    connection.is_active = True
                    db.session.commit()
                    
                    print(f"Token kaydedildi: {connection.auth_token}")  # Debug log
                    
                    return jsonify({
                        'success': True,
                        'message': 'Bağlantı başarılı',
                        'token': data['authToken']
                    })
                else:
                    print("API yanıtında authToken bulunamadı")  # Debug log
                    return jsonify({
                        'success': False,
                        'error': 'API yanıtında authToken bulunamadı'
                    }), 400
            except ValueError as e:
                print(f"JSON parse hatası: {str(e)}")  # Debug log
                return jsonify({
                    'success': False,
                    'error': 'API yanıtı geçerli bir JSON değil'
                }), 400
        else:
            print(f"API hata yanıtı: {response.status_code} - {response.text}")  # Debug log
            return jsonify({
                'success': False,
                'error': f'API yanıt kodu: {response.status_code}, Mesaj: {response.text}'
            }), response.status_code
            
    except requests.exceptions.ConnectionError as e:
        print(f"Bağlantı hatası: {str(e)}")  # Debug log
        return jsonify({
            'success': False,
            'error': 'Bağlantı hatası: Sunucuya ulaşılamıyor'
        }), 500
    except requests.exceptions.Timeout as e:
        print(f"Zaman aşımı hatası: {str(e)}")  # Debug log
        return jsonify({
            'success': False,
            'error': 'Bağlantı zaman aşımına uğradı'
        }), 500
    except Exception as e:
        print(f"Beklenmeyen hata: {str(e)}")  # Debug log
        return jsonify({
            'success': False,
            'error': f'Bir hata oluştu: {str(e)}'
        }), 500

@app.route('/rest-api/disconnect/<int:connection_id>', methods=['POST'])
@login_required
@csrf.exempt
def disconnect_rest_api_connection(connection_id):
    try:
        connection = RestApiConnection.query.get_or_404(connection_id)
        connection.is_active = False
        connection.auth_token = None
        db.session.commit()
        
        # Log ekle
        log = Log(
            log_type='system',
            level='INFO',
            message=f'REST API bağlantısı kesildi: {connection.name}',
            source='rest_api',
            created_by=current_user.id
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Bağlantı başarıyla kesildi'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/rest-api/scenario/<int:scenario_id>/toggle-status', methods=['POST', 'GET'])
@login_required
@csrf.exempt
def toggle_scenario_status_api(scenario_id):
    try:
        print(f"Senaryo ID: {scenario_id}")  # Debug log
        
        # Senaryoyu veritabanından al
        scenario = Scenario.query.get_or_404(scenario_id)
        if not scenario:
            print(f"Senaryo bulunamadı: {scenario_id}")  # Debug log
            return jsonify({
                'success': False,
                'message': f'Senaryo bulunamadı (ID: {scenario_id})'
            }), 404

        print(f"Değiştirilecek senaryo: {scenario.scenario_name}")  # Debug log
        
        # Aktif bağlantıları kontrol et
        active_connections = RestApiConnection.query.filter_by(is_active=True).all()
        if not active_connections:
            print("Aktif bağlantı bulunamadı")  # Debug log
            return jsonify({
                'success': False,
                'message': 'Aktif API bağlantısı bulunamadı'
            }), 400

        # Her aktif bağlantı için kontrol et
        for connection in active_connections:
            if not connection.auth_token:
                print(f"Bağlantı token'ı yok: {connection.name}")  # Debug log
                continue

            print(f"Bağlantı kontrol ediliyor: {connection.name}")  # Debug log

            # API isteği için headers hazırla
            headers = {
                'Authorization': f'EFTAdminAuthToken {connection.auth_token}',
                'Content-Type': 'application/json'
            }

            try:
                # Önce site ID'yi al
                site_url = f'{connection.base_url}/admin/v2/sites'
                print(f"Site bilgileri alınıyor: {site_url}")  # Debug log
                
                site_response = requests.get(
                    site_url,
                    headers=headers,
                    verify=False
                )

                print(f"Site API Yanıtı: {site_response.status_code} - {site_response.text}")  # Debug log

                if site_response.status_code != 200:
                    print(f"Site bilgileri alınamadı: {site_response.status_code}")  # Debug log
                    continue

                sites_data = site_response.json()
                if not sites_data.get('data'):
                    print("Site verisi bulunamadı")  # Debug log
                    continue

                site_id = sites_data['data'][0]['id']
                print(f"Site ID: {site_id}")  # Debug log

                # Event Rules'ları al
                rules_url = f'{connection.base_url}/admin/v2/sites/{site_id}/event-rules'
                print(f"Event Rules alınıyor: {rules_url}")  # Debug log
                
                response = requests.get(
                    rules_url,
                    headers=headers,
                    verify=False
                )

                print(f"Event Rules API Yanıtı: {response.status_code} - {response.text}")  # Debug log

                if response.status_code == 200:
                    event_rules = response.json()
                    if isinstance(event_rules, dict) and 'data' in event_rules:
                        event_rules = event_rules['data']
                        
                        print(f"Bulunan Event Rules sayısı: {len(event_rules)}")  # Debug log
                        
                        # Senaryo adına göre eşleşen kuralı bul
                        matching_rule = None
                        for rule in event_rules:
                            rule_name = rule.get('attributes', {}).get('info', {}).get('Name')
                            print(f"Kontrol edilen kural: {rule_name}")  # Debug log
                            if rule_name == scenario.scenario_name:
                                matching_rule = rule
                                print(f"Eşleşen kural bulundu: {rule_name}")  # Debug log
                                break

                        if matching_rule:
                            rule_id = matching_rule['id']
                            current_enabled = matching_rule.get('attributes', {}).get('info', {}).get('Enabled', True)
                            new_enabled = not current_enabled

                            # Eğer zaten istenen durumdaysa, API'ye istek göndermeden başarılı dön
                            if current_enabled == new_enabled:
                                scenario.status = 'active' if new_enabled else 'inactive'
                                db.session.commit()
                                return jsonify({
                                    'success': True,
                                    'message': f'Senaryo zaten {"aktif" if new_enabled else "pasif"}',
                                    'new_status': 'active' if new_enabled else 'inactive'
                                })

                            print(f"Kural durumu değiştiriliyor: {rule_id} - {current_enabled} -> {new_enabled}")  # Debug log

                            update_url = f'{connection.base_url}/admin/v2/sites/{site_id}/event-rules/{rule_id}'
                            update_payload = {
                                "data": {
                                    "type": "event-rule",
                                    "id": rule_id,
                                    "attributes": {
                                        "info": {
                                            "Enabled": new_enabled
                                        }
                                    }
                                }
                            }

                            print(f"Güncelleme isteği gönderiliyor: {update_url}")  # Debug log
                            print(f"Güncelleme payload: {update_payload}")  # Debug log

                            update_response = requests.patch(
                                update_url,
                                headers=headers,
                                json=update_payload,
                                verify=False
                            )

                            print(f"Güncelleme yanıtı: {update_response.status_code} - {update_response.text}")  # Debug log

                            if update_response.status_code in [200, 201]:
                                scenario.status = 'active' if new_enabled else 'inactive'
                                db.session.commit()
                                return jsonify({
                                    'success': True,
                                    'message': f'Senaryo durumu başarıyla güncellendi: {"aktif" if new_enabled else "pasif"}',
                                    'new_status': 'active' if new_enabled else 'inactive'
                                })
                            elif update_response.status_code == 401:
                                return jsonify({
                                    'success': False,
                                    'message': 'API token süresi dolmuş. Lütfen bağlantıyı yenileyin.'
                                }), 401
                            elif update_response.status_code == 400:
                                return jsonify({
                                    'success': False,
                                    'message': 'Aktif Token bağlantısı bulunmadı.'
                                }), 400
                            elif update_response.status_code == 404:
                                return jsonify({
                                    'success': False,
                                    'message': 'Token süresi dolmuş veya bağlantı bulunamadı. Lütfen bağlantıyı kontrol edin.'
                                }), 404
                            else:
                                return jsonify({
                                    'success': False,
                                    'message': f'API güncelleme hatası: {update_response.text}'
                                }), update_response.status_code
                        else:
                            print(f"Eşleşen kural bulunamadı: {scenario.scenario_name}")  # Debug log
                            return jsonify({
                                'success': False,
                                'message': f'API\'de "{scenario.scenario_name}" adlı senaryo bulunamadı'
                            }), 404

            except Exception as e:
                print(f"Bağlantı hatası ({connection.name}): {str(e)}")  # Debug log
                continue

        # Hiçbir bağlantıda eşleşen senaryo bulunamadı
        return jsonify({
            'success': False,
            'message': f'API\'de "{scenario.scenario_name}" adlı senaryo bulunamadı'
        }), 404

    except Exception as e:
        print(f"Genel hata: {str(e)}")  # Debug log
        return jsonify({
            'success': False,
            'message': f'Senaryo durumu güncellenirken bir hata oluştu: {str(e)}'
        }), 500

@app.route('/event-rules/<int:connection_id>')
@login_required
def event_rules(connection_id):
    try:
        # Bağlantıyı kontrol et
        connection = RestApiConnection.query.get_or_404(connection_id)
        if not connection.is_active or not connection.auth_token:
            flash('Bağlantı aktif değil veya geçerli bir token yok', 'error')
            return redirect(url_for('rest_api_connections'))

        # API isteği için headers hazırla
        headers = {
            'Authorization': f'EFTAdminAuthToken {connection.auth_token}',
            'Content-Type': 'application/json'
        }

        # Önce site ID'yi al
        site_response = requests.get(
            f'{connection.base_url}/admin/v2/sites',
            headers=headers,
            verify=False
        )

        print(f"Site API Response Status: {site_response.status_code}")  # Debug log
        print(f"Site API Response Text: {site_response.text}")  # Debug log

        if site_response.status_code != 200:
            flash(f'Site bilgileri alınamadı: {site_response.status_code} - {site_response.text}', 'error')
            return redirect(url_for('rest_api_connections'))

        try:
            sites_data = site_response.json()
            if not sites_data.get('data'):
                flash('Site bilgisi bulunamadı', 'error')
                return redirect(url_for('rest_api_connections'))

            site_id = sites_data['data'][0]['id']  # İlk siteyi al

            # Event Rules'ları al
            response = requests.get(
                f'{connection.base_url}/admin/v2/sites/{site_id}/event-rules',
                headers=headers,
                verify=False
            )

            print(f"Event Rules API URL: {connection.base_url}/admin/v2/sites/{site_id}/event-rules")  # Debug log
            print(f"Event Rules API Headers: {headers}")  # Debug log
            print(f"Event Rules API Response Status: {response.status_code}")  # Debug log
            print(f"Event Rules API Response Text: {response.text}")  # Debug log

            if response.status_code == 200:
                try:
                    event_rules = response.json()
                    # API yanıtını kontrol et
                    if isinstance(event_rules, dict) and 'data' in event_rules:
                        event_rules = event_rules['data']
                    return render_template('event_rules.html', 
                                            connection=connection, 
                                            event_rules=event_rules)
                except ValueError as e:
                    flash(f'API yanıtı geçersiz JSON formatında: {str(e)}', 'error')
                    return redirect(url_for('rest_api_connections'))
            else:
                flash(f'Event Rules alınırken bir hata oluştu: {response.status_code} - {response.text}', 'error')
                return redirect(url_for('rest_api_connections'))

        except (KeyError, IndexError) as e:
            flash(f'Site bilgileri işlenirken bir hata oluştu: {str(e)}', 'error')
            return redirect(url_for('rest_api_connections'))

    except Exception as e:
        flash(f'Event Rules alınırken bir hata oluştu: {str(e)}', 'error')
        return redirect(url_for('rest_api_connections'))

@app.route('/rest-api/event-rules/<int:connection_id>/<rule_id>/toggle', methods=['POST'])
@login_required
@csrf.exempt
def toggle_event_rule(connection_id, rule_id):
    try:
        # Bağlantıyı kontrol et
        connection = RestApiConnection.query.get_or_404(connection_id)
        if not connection.is_active or not connection.auth_token:
            return jsonify({
                'success': False,
                'error': 'Bağlantı aktif değil veya geçerli bir token yok'
            }), 400

        # API isteği için gerekli bilgileri hazırla
        headers = {
            'Authorization': f'EFTAdminAuthToken {connection.auth_token}',
            'Content-Type': 'application/json'
        }
        
        # Önce mevcut durumu al
        api_url = f'http://{connection.ip_address}:{connection.port}/admin/v2/sites/1/event-rules/{rule_id}'
        response = requests.get(
            api_url,
            headers=headers,
            verify=False
        )
        
        if response.status_code != 200:
            return jsonify({
                'success': False,
                'error': f'API yanıt hatası: {response.status_code} - {response.text}'
            }), response.status_code

        current_status = response.json().get('status')
        new_status = 'inactive' if current_status == 'active' else 'active'
        
        # Durumu güncelle
        response = requests.put(
            api_url,
            headers=headers,
            json={'status': new_status},
            verify=False
        )
        
        if response.status_code == 200:
            # Log ekle
            log = Log(
                log_type='system',
                level='INFO',
                message=f'Event Rule durumu değiştirildi: {rule_id} - {new_status}',
                source='rest_api',
                created_by=current_user.id
            )
            db.session.add(log)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Event Rule durumu başarıyla güncellendi',
                'new_status': new_status
            })
        else:
            return jsonify({
                'success': False,
                'error': f'API yanıt hatası: {response.status_code} - {response.text}'
            }), response.status_code

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/rest-api/event-rules/<int:connection_id>/<rule_id>/update', methods=['POST'])
@login_required
@csrf.exempt
def update_event_rule(connection_id, rule_id):
    try:
        # Bağlantıyı kontrol et
        connection = RestApiConnection.query.get_or_404(connection_id)
        if not connection.is_active or not connection.auth_token:
            return jsonify({
                'success': False,
                'error': 'Bağlantı aktif değil veya geçerli bir token yok'
            }), 400

        # API isteği için headers hazırla
        headers = {
            'Authorization': f'EFTAdminAuthToken {connection.auth_token}',
            'Content-Type': 'application/json'
        }

        # Request body'den verileri al
        data = request.get_json()
        if not data or 'name' not in data or 'description' not in data:
            return jsonify({
                'success': False,
                'error': 'Name ve description alanları zorunludur'
            }), 400

        # API isteği için body hazırla
        payload = {
            'data': {
                'type': 'eventRule',
                'attributes': {
                    'info': {
                        'Name': data['name'],
                        'Description': data['description']
                    }
                }
            }
        }

        # API'ye PATCH isteği gönder
        response = requests.patch(
            f'{connection.base_url}/admin/v2/sites/1/event-rules/{rule_id}',
            headers=headers,
            json=payload,
            verify=False
        )

        if response.status_code == 200:
            # Log kaydı oluştur
            log = Log(
                log_type='manual',
                level='INFO',
                message=f'Event Rule güncellendi: {data["name"]}',
                source='REST API',
                created_by=current_user.id
            )
            db.session.add(log)
            db.session.commit()

            return jsonify({
                'success': True,
                'message': 'Event Rule başarıyla güncellendi'
            })
        else:
            return jsonify({
                'success': False,
                'error': f'API yanıt hatası: {response.status_code} - {response.text}'
            }), 400

    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Hata oluştu: {str(e)}'
        }), 500

@app.route('/event-rules/<int:connection_id>/<string:rule_id>')
@login_required
def event_rule_detail(connection_id, rule_id):
    connection = RestApiConnection.query.get_or_404(connection_id)
    
    # Get event rule details from API
    response = requests.get(
        f'{connection.base_url}/admin/v2/sites/1/eventrules/{rule_id}',
        headers={'Authorization': f'Bearer {connection.token}'},
        verify=False
    )
    
    if response.status_code == 200:
        rule = response.json()
        return render_template('event_rule_detail.html', connection=connection, rule=rule)
    else:
        flash('Event rule detayları alınamadı.', 'error')
        return redirect(url_for('event_rules', connection_id=connection_id))

# Error handlers for better debugging
@app.errorhandler(500)
def internal_error(error):
    """Handle 500 Internal Server Error"""
    print(f"Internal Server Error: {error}")
    return render_template('error.html', 
                         error_code=500, 
                         error_message="Internal Server Error. Please check the logs."), 500

@app.errorhandler(404)
def not_found_error(error):
    """Handle 404 Not Found Error"""
    return render_template('error.html', 
                         error_code=404, 
                         error_message="Page not found."), 404

if __name__ == '__main__':
    with app.app_context():
        initialize_permissions()
    
    # Azure App Service port configuration
    port = int(os.environ.get('PORT', 5050))
    
    # Check if running in Azure (production)
    if os.environ.get('WEBSITE_SITE_NAME'):
        # Running in Azure App Service
        app.run(
            host='0.0.0.0',
            port=port,
            debug=False
        )
    else:
        # Running locally
        app.run(
            host='0.0.0.0',
            port=port,
            ssl_context=('cert.pem', 'key.pem'),
            debug=True
        )